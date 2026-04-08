import os
import csv
import history_manager
import configparser
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import colorama
from colorama import Fore, Style, init
import threading
from rich.console import Console
from rich.table import Table
from rich import box

_console = Console(highlight=False, legacy_windows=False)

def _fmt_bool(v):
    return "[green]Ja[/]" if v else "[dim]Nein[/]"
from utils import safe_read_config

# Colorama initialisieren
init(autoreset=True)

# Thread-sicherer Zugriff
console_lock = threading.Lock()

def thread_safe_print(color, message):
    with console_lock:
        print(f"{color}{message}{Style.RESET_ALL}", flush=True)

# Hilfsfunktionen
def print_error(message):
    thread_safe_print(Fore.RED, f"❌ {message}")

def print_warning(message):
    thread_safe_print(Fore.YELLOW, f"⚠️ {message}")

def print_admin_warning(message):
    thread_safe_print(Fore.LIGHTRED_EX, f"❗ {message}")

def print_warningtext(message):
    thread_safe_print(Fore.MAGENTA, f"👾 {message}")

def print_success(message):
    thread_safe_print(Fore.GREEN, f"✅ {message}")

def print_info(message):
    thread_safe_print(Fore.CYAN, f"ℹ️ {message}")

def print_creation(message):
    thread_safe_print(Fore.WHITE, f"✨ {message}")

def print_section(title, count=None):
    """Gibt eine Abschnittsüberschrift aus – strukturelle Trennung, kein Emoji."""
    label = f"{title} ({count})" if count is not None else title
    _console.rule(label, style="cyan")

# --- Validierungskonstanten ---
SCHILD_REQUIRED_COLUMNS = [
    'Interne ID-Nummer', 'Nachname', 'Vorname', 'Klasse', 'Klassenlehrer', 
    'Geburtsdatum', 'Geschlecht', 'vorauss. Abschlussdatum', 'Aufnahmedatum', 
    'Entlassdatum', 'Volljährig', 'Schulpflicht erfüllt', 'Status'
]
SCHILD_OPTIONAL_COLUMNS = {
    'E-Mail (privat)': 'E-Mail-Korrespondenz mit Schülern über WebUntis nicht möglich.',
    'Telefon-Nr.': 'Notfall-Kontaktdaten (Telefon) fehlen in WebUntis.',
    'Fax-Nr.': 'Faxnummern fehlen in WebUntis.',
    'Straße': 'Adressdaten (Straße) der Schüler fehlen in WebUntis.',
    'Postleitzahl': 'Adressdaten (PLZ) der Schüler fehlen in WebUntis.',
    'Ortsname': 'Adressdaten (Ort) der Schüler fehlen in WebUntis.'
}

TEACHER_REQUIRED_COLUMNS = [
    'name', 'longName', 'foreName', 'title', 'birthDate', 'pnr', 
    'address.email', 'address.phone', 'address.mobile', 'address.street', 
    'address.postCode', 'address.city'
]

CLASS_REQUIRED_COLUMNS = [
    'Auswahl', '', 'Klasse', 'Langname', 'Alias', 'Jahrgangsstufe',
    'Text', 'Klassenlehrkraft', 'Klassenlehrkraft', 'Abteilung', 'Von', 'Bis'
]

# Werte die als semantisch gleichwertig zu "leer" behandelt werden sollen.
# Wechsel zwischen diesen Werten und "" wird nicht als Änderung gewertet.
_EMPTY_EQUIVALENTS = {'', 'nein', 'keine', 'keiner', 'keines', 'kein', 'false'}

def normalize_value(value):
    """Normalisiert einen Feldwert für den Vergleich.
    Werte wie 'Nein', 'Keine' etc. werden wie leer behandelt."""
    v = value.strip().replace('\u00a0', '').lower()
    if v in _EMPTY_EQUIVALENTS:
        return ''
    return value.strip().replace('\u00a0', '')

def validate_imports():
    """
    Führt einen Vorab-Check aller benötigten Export-Dateien durch.
    Gibt einen Report über fehlende Dateien oder Spalten zurück.
    """
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    
    report = {
        "success": True,
        "files": {}
    }

    # 1. Schild-Export prüfen
    schild_dir = config.get('Directories', 'schildexport_directory', fallback='.')
    if schild_dir in ('.', '', None): schild_dir = os.getcwd()
    
    schild_files = [f for f in os.listdir(schild_dir) if f.lower().endswith('.csv')]
    if not schild_files:
        report["files"]["Schild-Export"] = {"status": "error", "message": "Keine CSV-Datei im Schild-Export Verzeichnis gefunden."}
        report["success"] = False
    else:
        newest_schild = max(schild_files, key=lambda f: os.path.getctime(os.path.join(schild_dir, f)))
        file_path = os.path.join(schild_dir, newest_schild)
        try:
            with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                # Erstes Zeichen prüfen für Separator-Check
                first_line = f.readline()
                f.seek(0)
                
                separator = ';'
                if ';' not in first_line and ',' in first_line:
                    report["files"]["Schild-Export"] = {"status": "error", "message": f"Falscher Separator in '{newest_schild}'. Erwartet: ';', gefunden: ','"}
                    report["success"] = False
                else:
                    reader = csv.DictReader(f, delimiter=';')
                    headers = reader.fieldnames or []
                    
                    missing_req = [col for col in SCHILD_REQUIRED_COLUMNS if col not in headers]
                    missing_opt = [col for col in SCHILD_OPTIONAL_COLUMNS if col not in headers]
                    
                    file_status = "success"
                    messages = []
                    
                    if missing_req:
                        file_status = "error"
                        messages.append(f"Erforderliche Spalten fehlen: {', '.join(missing_req)}")
                        report["success"] = False
                    
                    if missing_opt:
                        if file_status != "error": file_status = "warning"
                        for col in missing_opt:
                            messages.append(f"Info (Optional): Spalte '{col}' fehlt. ({SCHILD_OPTIONAL_COLUMNS[col]})")
                    
                    report["files"]["Schild-Export"] = {
                        "status": file_status, 
                        "file": newest_schild,
                        "messages": messages
                    }
        except Exception as e:
            report["files"]["Schild-Export"] = {"status": "error", "message": f"Fehler beim Lesen: {str(e)}"}
            report["success"] = False

    # 2. Lehrer-Daten prüfen
    teachers_dir = config.get('Directories', 'teachers_directory', fallback='./Lehrerdaten')
    if not os.path.exists(teachers_dir):
        report["files"]["Lehrer-Daten"] = {"status": "error", "message": f"Verzeichnis '{teachers_dir}' nicht gefunden."}
        report["success"] = False
    else:
        teacher_files = [f for f in os.listdir(teachers_dir) if f.lower().endswith('.csv')]
        if not teacher_files:
            report["files"]["Lehrer-Daten"] = {"status": "error", "message": "Keine CSV-Datei im Lehrerdatenverzeichnis gefunden."}
            report["success"] = False
        else:
            newest_teacher = max(teacher_files, key=lambda f: os.path.getctime(os.path.join(teachers_dir, f)))
            file_path = os.path.join(teachers_dir, newest_teacher)
            try:
                with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                    reader = csv.DictReader(f, delimiter='\t')
                    headers = reader.fieldnames or []
                    missing = [col for col in TEACHER_REQUIRED_COLUMNS if col not in headers]
                    
                    if missing:
                        report["files"]["Lehrer-Daten"] = {"status": "error", "file": newest_teacher, "message": f"Erforderliche Spalten fehlen (Tab-Trenner prüfen!): {', '.join(missing)}"}
                        report["success"] = False
                    else:
                        report["files"]["Lehrer-Daten"] = {"status": "success", "file": newest_teacher}
            except Exception as e:
                report["files"]["Lehrer-Daten"] = {"status": "error", "message": f"Fehler beim Lesen: {str(e)}"}
                report["success"] = False

    # 3. Klassen-Daten prüfen
    classes_dir = config.get('Directories', 'classes_directory', fallback='./Klassendaten')
    if not os.path.exists(classes_dir):
        report["files"]["Klassen-Daten"] = {"status": "error", "message": f"Verzeichnis '{classes_dir}' nicht gefunden."}
        report["success"] = False
    else:
        class_files = [f for f in os.listdir(classes_dir) if f.lower().endswith('.csv')]
        if not class_files:
            report["files"]["Klassen-Daten"] = {"status": "error", "message": "Keine CSV-Datei im Klassendatenverzeichnis gefunden."}
            report["success"] = False
        else:
            newest_class = max(class_files, key=lambda f: os.path.getctime(os.path.join(classes_dir, f)))
            file_path = os.path.join(classes_dir, newest_class)
            try:
                with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                    reader = csv.reader(f, delimiter=';')
                    headers = next(reader, [])
                    
                    # Klassen-Export hat oft exakte Spaltenreihenfolge
                    if len(headers) < len(CLASS_REQUIRED_COLUMNS):
                        report["files"]["Klassen-Daten"] = {"status": "error", "file": newest_class, "message": f"Zu wenige Spalten gefunden ({len(headers)} statt {len(CLASS_REQUIRED_COLUMNS)})."}
                        report["success"] = False
                    else:
                        # Wir prüfen die ersten 12 Spalten
                        # Die 2. Spalte ist oft ein leerer String oder '[eine Leere Spalte]' im README
                        # Wir prüfen hier eher die Existenz kritischer Spalten wie 'Klasse', 'Langname', 'Klassenlehrkraft'
                        critical_checks = {2: 'Klasse', 3: 'Langname', 7: 'Klassenlehrkraft'}
                        missing_crit = [v for k, v in critical_checks.items() if len(headers) <= k or headers[k] != v]
                        
                        if missing_crit:
                            report["files"]["Klassen-Daten"] = {"status": "error", "file": newest_class, "message": f"Kritische Spaltennamen falsch oder an falscher Position: {', '.join(missing_crit)}"}
                            report["success"] = False
                        else:
                            report["files"]["Klassen-Daten"] = {"status": "success", "file": newest_class}
            except Exception as e:
                report["files"]["Klassen-Daten"] = {"status": "error", "message": f"Fehler beim Lesen: {str(e)}"}
                report["success"] = False

    return report

def run(use_abschlussdatum=False, create_second_file=False, enable_attestpflicht_column=False, create_class_size_file=False, disable_import_file_creation=False, disable_import_file_if_admin_warning=False, warn_entlassdatum=True, warn_aufnahmedatum=True, warn_klassenwechsel=True, warn_new_students=True, warn_karteileichen=False, class_change_recipients="both", no_log=False, no_xlsx=False, admin_warnings_cache=None, enable_nachteilsausgleich_column=False):
    if admin_warnings_cache is None:
        admin_warnings_cache = []
    # Hauptfunktion zur Verarbeitung der Daten und Generierung von Warnungen
    print_section("Hauptverarbeitung")

    warnings = []
    class_change_warnings = []
    admission_date_warnings = []
    new_student_warnings = []
    karteileichen_warnings = []

    # Konfigurationsdatei einlesen
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    classes_dir  = config.get('Directories', 'classes_directory')
    teachers_dir = config.get('Directories', 'teachers_directory')

    opts = Table(box=None, show_header=False, pad_edge=False, padding=(0, 2, 0, 2))
    opts.add_column(style="cyan dim", no_wrap=True)
    opts.add_column()
    opts.add_row("Abschlussdatum",               _fmt_bool(use_abschlussdatum))
    opts.add_row("Zweite Importdatei",            _fmt_bool(create_second_file))
    opts.add_row("Klassengrößen-Auswertung",      _fmt_bool(create_class_size_file))
    opts.add_row("Attestpflicht-Spalte",          _fmt_bool(enable_attestpflicht_column))
    opts.add_row("Nachteilsausgleich-Spalte",     _fmt_bool(enable_nachteilsausgleich_column))
    opts.add_row("Importdatei unterdrücken",      _fmt_bool(disable_import_file_creation))
    opts.add_row("Unterdrücken bei Admin-Warnung",_fmt_bool(disable_import_file_if_admin_warning))
    opts.add_row("Warnung: Entlassdatum",         _fmt_bool(warn_entlassdatum))
    opts.add_row("Warnung: Aufnahmedatum",        _fmt_bool(warn_aufnahmedatum))
    opts.add_row("Warnung: Klassenwechsel",       _fmt_bool(warn_klassenwechsel))
    opts.add_row("Warnung: Neue Schüler",         _fmt_bool(warn_new_students))
    opts.add_row("Warnung: Karteileichen",        _fmt_bool(warn_karteileichen))
    opts.add_row("Text-Logs",                     _fmt_bool(not no_log))
    opts.add_row("Excel-Logs",                    _fmt_bool(not no_xlsx))
    opts.add_row("KW E-Mail-Empfänger",           str(class_change_recipients) if class_change_recipients else "[dim]–[/]")
    opts.add_row("Klassen-CSV",                   classes_dir)
    opts.add_row("Lehrkräfte-CSV",                teachers_dir)
    _console.print(opts)

    # Daten einlesen und verarbeiten
    classes_by_name = read_classes(classes_dir, teachers_dir)
    output_data_students, students_by_id = read_students(use_abschlussdatum)

    # Historische Klassen-Zuordnung synchronisieren
    sync_student_classes_from_latest()

    # Warnungen erstellen basierend auf der Benutzerauswahl
    if warn_entlassdatum:
        warnings = create_warnings(classes_by_name, students_by_id)
    if warn_klassenwechsel:
        class_change_warnings = create_class_change_warnings(classes_by_name, students_by_id, class_change_recipients)
    if warn_aufnahmedatum:
        admission_date_warnings = create_admission_date_warnings(classes_by_name, students_by_id)
    if warn_new_students:
        new_student_warnings = create_new_student_warnings(classes_by_name, students_by_id)
    if warn_karteileichen:
        karteileichen_warnings = create_karteileichen_warnings(classes_by_name, students_by_id, admin_warnings_cache)

    # Alle Warnungen zusammenführen
    all_warnings = warnings + class_change_warnings + admission_date_warnings + new_student_warnings + karteileichen_warnings

    # Separate Konsolenausgabe der verschiedenen Warnungen
    print_section("Entlassdatum-Warnungen",   len(warnings))
    print_warnings(warnings)
    print_section("Klassenwechsel-Warnungen",  len(class_change_warnings))
    print_warnings(class_change_warnings)
    print_section("Aufnahmedatum-Warnungen",   len(admission_date_warnings))
    print_warnings(admission_date_warnings)
    print_section("Neue-Schüler-Warnungen",    len(new_student_warnings))
    print_warnings(new_student_warnings)
    print_section("Karteileichen-Warnungen",   len(karteileichen_warnings))
    print_warnings(karteileichen_warnings)

    # Dateien speichern
    print_section("Ausgabe")
    save_files(output_data_students, all_warnings, create_second_file, admin_warnings_cache, disable_import_file_creation, disable_import_file_if_admin_warning, enable_attestpflicht_column, enable_nachteilsausgleich_column)

    if create_class_size_file:
     create_class_sizes_file(students_by_id)

    # Vergleiche die letzten beiden Dateien und erfasse Feldänderungen für Info-Mails
    info_changes = compare_latest_imports(no_log=no_log, no_xlsx=no_xlsx) or []

    print_success("Hauptverarbeitung abgeschlossen.")
    return all_warnings, info_changes

INFO_MAIL_AVAILABLE_FIELDS = [
    'Vorname', 'Nachname', 'Geschlecht', 'Geburtsdatum',
    'Entlassdatum', 'Aufnahmedatum', 'vorauss. Abschlussdatum',
    'Schulpflicht', 'Volljährig', 'Aktiv',
    'Attestpflicht', 'Nachteilsausgleich',
]

def create_info_notifications(changes, selected_fields):
    """Erstellt Info-Mail-Objekte für Schüler mit Änderungen in den gewählten Feldern."""
    if not changes or not selected_fields:
        return []

    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    classes_dir  = config.get('Directories', 'classes_directory', fallback='Klassendaten')
    teachers_dir = config.get('Directories', 'teachers_directory', fallback='Lehrerdaten')
    classes_by_name = read_classes(classes_dir, teachers_dir)

    notifications = []
    for change in changes:
        relevant = {k: v for k, v in change['changes'].items() if k in selected_fields}
        if not relevant:
            continue

        current_class = change['current_class']
        class_info    = classes_by_name.get(current_class, {})
        student       = change.get('current_student', {})
        vorname  = student.get('Vorname',  '') or change['name'].split(' ', 1)[0]
        nachname = student.get('Nachname', '') or (change['name'].split(' ', 1)[1] if ' ' in change['name'] else '')

        rows = ''.join(
            f'<tr>'
            f'<td style="padding:4px 8px;border:1px solid #ccc"><strong>{field}</strong></td>'
            f'<td style="padding:4px 8px;border:1px solid #ccc;color:#c00">{vals["old"]}</td>'
            f'<td style="padding:4px 8px;text-align:center">&#8594;</td>'
            f'<td style="padding:4px 8px;border:1px solid #ccc;color:#060">{vals["new"]}</td>'
            f'</tr>'
            for field, vals in relevant.items()
        )
        aenderungen_html = (
            '<table style="border-collapse:collapse;width:100%">'
            '<thead><tr>'
            '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0">Feld</th>'
            '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0">Alt</th>'
            '<th style="padding:4px 8px;background:#f0f0f0"></th>'
            '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0">Neu</th>'
            f'</tr></thead><tbody>{rows}</tbody></table>'
        )

        notifications.append({
            'Vorname':               vorname,
            'Nachname':              nachname,
            'Klasse':                current_class,
            'aenderungen_html':      aenderungen_html,
            'aenderungen_text':      '\n'.join(f'{k}: {v["old"]} -> {v["new"]}' for k, v in relevant.items()),
            'aenderungen_felder':    ', '.join(relevant.keys()),
            'Klassenlehrkraft_1':       class_info.get('Klassenlehrkraft_1',       'N/A'),
            'Klassenlehrkraft_1_Email': class_info.get('Klassenlehrkraft_1_Email', 'N/A'),
            'Klassenlehrkraft_2':       class_info.get('Klassenlehrkraft_2',       'N/A'),
            'Klassenlehrkraft_2_Email': class_info.get('Klassenlehrkraft_2_Email', 'N/A'),
            'info_notification': True,
            'status': 'offen',
        })

    return notifications

def get_directory(key, default=None):
    # Hilfsfunktion zum Abrufen von Verzeichnispfaden aus der Konfigurationsdatei
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    return config.get('Directories', key, fallback=default)

def sync_student_classes_from_latest():
    """Liest den neuesten Schild-Export und aktualisiert Namen und Klassen in der Historien-DB."""
    print_info("Synchronisiere Schüler-Klassen-Mapping mit der Historien-Datenbank...")
    try:
        # 1. Neuesten Import finden (WebUntis Importe Verzeichnis)
        import_dir = get_directory('import_directory', './WebUntis Importe')
        if not os.path.exists(import_dir):
            print_warning(f"Verzeichnis {import_dir} existiert nicht.")
            return False

        output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
        if not output_files:
            print_warning("Keine Importdatei zur Synchronisierung gefunden.")
            return False
            
        newest_file = max(output_files, key=lambda f: os.path.getctime(os.path.join(import_dir, f)))
        full_path = os.path.join(import_dir, newest_file)
        data = read_csv(full_path)
        
        # 2. Mapping erstellen
        student_list = []
        for student_id, row in data.items():
            student_list.append({
                'id': student_id,
                'name': f"{row.get('Vorname', '')} {row.get('Nachname', '')}",
                'class': row.get('Klasse', '')
            })
            
        # 3. Bulk Update in history_manager
        if student_list:
            history_manager.bulk_update_students(student_list)
            return True
    except Exception as e:
        print_error(f"Fehler bei der Klassen-Synchronisierung: {e}")
    return False


def read_csv(file_path):
    # Funktion zum Einlesen einer CSV-Datei und Rückgabe eines Dictionaries mit den Daten
    data = {}
    try:
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                # Entfernt unsichtbare Zeichen und Leerzeichen
                cleaned_row = {key: value.strip().replace('\u00a0', '') for key, value in row.items()}
                student_id = cleaned_row.get('Interne ID-Nummer')
                if student_id:
                    data[student_id] = cleaned_row
    except Exception as e:
        print_error(f"Fehler beim Lesen der Datei {file_path}: {e}")
    return data


def compare_latest_imports(no_log=False, no_xlsx=False):
    # Funktion zum Vergleichen der neuesten Importdateien und Erfassen von Änderungen
    print_section("Log-Erstellung")
    # Überspringen der Erstellung je nach Flags
    if no_log:
        print_info("Log-Datei-Erstellung wurde deaktiviert.")
    if no_xlsx:
        print_info("Excel-Datei-Erstellung wurde deaktiviert.")
    # Abbruch, wenn sowohl no_log als auch no_xlsx aktiviert sind
    if no_log and no_xlsx:
        print_warning("Vergleich abgebrochen: Sowohl Log- als auch Excel-Datei-Erstellung wurden deaktiviert.")
        return []

    # Verzeichnisse definieren
    import_dir = get_directory('import_directory', './WebUntis Importe')
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")

    # Verzeichnisse für Log- und Excel-Dateien aus der settings.ini lesen
    log_dir = config.get("Directories", "log_directory", fallback="./Logs")
    xlsx_dir = config.get("Directories", "xlsx_directory", fallback="ExcelLogs")

    # Sicherstellen, dass die Verzeichnisse existieren
    os.makedirs(import_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    # Neueste zwei Dateien im Importverzeichnis finden
    csv_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    if len(csv_files) < 2:
        print_warning(f"Nicht genügend Dateien im Verzeichnis {import_dir} vorhanden, um einen Vergleich durchzuführen. Abbruch der Log-Erstellung.")
        return []

    # Nach Erstellungszeit sortieren und die letzten beiden Dateien auswählen
    csv_files.sort(key=lambda f: os.path.getctime(os.path.join(import_dir, f)), reverse=True)
    latest_file = csv_files[0]
    previous_file = csv_files[1]

    print_info(f"  Vergleiche: '{previous_file}' → '{latest_file}'")

    # Dateien einlesen
    previous_data = read_csv(os.path.join(import_dir, previous_file))
    latest_data = read_csv(os.path.join(import_dir, latest_file))

    # Änderungen erfassen
    changes = []
    for student_id, latest_student in latest_data.items():
        previous_student = previous_data.get(student_id)
        if not previous_student:
            continue  # Neuer Schüler, keine Änderung

        row_changed = False
        updated_row = latest_student.copy()
        change_details = {}
        for field, new_value in latest_student.items():
            old_raw = previous_student.get(field, '').strip().replace('\u00a0', '')
            new_raw = new_value.strip().replace('\u00a0', '')

            if normalize_value(old_raw) != normalize_value(new_raw):
                row_changed = True
                updated_row[field] = f"{old_raw} -> {new_raw}"
                change_details[field] = {"old": old_raw, "new": new_raw}

        if row_changed:
            changes.append({
                "student_id": student_id,
                "name": f"{latest_student.get('Vorname', '')} {latest_student.get('Nachname', '')}",
                "current_class": latest_student.get('Klasse', ''),
                "changes": change_details,
                "row": updated_row,
                "current_student": latest_student.copy(),
            })
    if not changes:
        print_warning("Keine Änderungen zwischen den Dateien festgestellt. Abbruch der Log-Erstellung.")
        print_warningtext("Prüfen Sie, ob Sie die selbe Schild-Export Datei zwei mal verarbeitet haben.")
        return []

    # Datum und Uhrzeit für Dateinamen
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file_path = os.path.join(log_dir, f'ÄnderungsLog_{timestamp}.log')
    excel_file_path = os.path.join(xlsx_dir, f'ÄnderungsLog_{timestamp}.xlsx')
    
    # Änderungen in der Historien-Datenbank speichern
    history_manager.record_changes(changes, timestamp, previous_file, latest_file)

    # Log-Datei erstellen
    if not no_log:
        print_creation(f"Erstelle Text-Log-Datei: {log_file_path}")
        with open(log_file_path, 'w', encoding='utf-8') as log_file:
            for change in changes:
                log_file.write(f"Schüler: {change['name']} (ID: {change['student_id']}) [Klasse: {change['current_class']}]\n")
                for field, values in change['changes'].items():
                    log_file.write(f"  {field}: {values['old']} -> {values['new']}\n")
                log_file.write("\n")
        print_success(".log Log-Datei erfolgreich erstellt.")

    # Excel-Datei erstellen
    if not no_xlsx and changes:
        print_creation(f"Erstelle Excel-Log-Datei: {excel_file_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Änderungen"

        # Kopfzeile schreiben
        headers = list(changes[0]["row"].keys())
        ws.append(headers)

        # Daten schreiben
        for change in changes:
            row = []
            for header in headers:
                value = change["row"].get(header, "")
                row.append(value)
            ws.append(row)

        # Spaltenbreite automatisch anpassen
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Hol den Buchstaben der Spalte
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Änderungen farblich hervorheben
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if "->" in str(cell.value):
                    cell.fill = red_fill

        wb.save(excel_file_path)
        print_success("Excel-Log-Datei erfolgreich erstellt.")
    print_success("Vergleich der neuesten Importdateien abgeschlossen.")
    return changes

from smtp import send_email
def compare_timeframe_imports(timeframe_hours=24, no_log=False, no_xlsx=False):
    # Funktion zum Vergleichen von Importdateien innerhalb eines bestimmten Zeitrahmens und Senden von E-Mails bei Änderungen
    print_section("Zeitrahmen-Vergleich")

    # Verzeichnisse und Einstellungen
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    email_config = configparser.ConfigParser()
    email_config.read("email_settings.ini", encoding='utf-8-sig')

    timeframe_hours = config.getint("ProcessingOptions", "timeframe_hours", fallback=24)
    import_dir = config.get("Directories", "import_directory", fallback="./WebUntis Importe")
    log_dir = config.get("Directories", "log_directory", fallback="./Logs")
    xlsx_dir = config.get("Directories", "xlsx_directory", fallback="ExcelLogs")
    admin_email = email_config.get("Email", "admin_email", fallback=None)

    if not admin_email:
        print_warning("Keine Admin-E-Mail-Adresse in der email_settings.ini definiert. E-Mail-Versand deaktiviert.")
        send_log_email = False

    os.makedirs(import_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    # Überprüfung des letzten E-Mail-Versands
    print_info(f"Überprüfe, ob in den letzten {timeframe_hours} Stunden schon eine Log-E-Mail versendet wurde...")
    email_log_path = os.path.join(log_dir, "last_email_sent.log")
    if os.path.exists(email_log_path):
        with open(email_log_path, "r") as email_log:
            last_email_time = datetime.strptime(email_log.read().strip(), "%Y-%m-%d %H:%M:%S")
        if datetime.now() - last_email_time < timedelta(hours=timeframe_hours):
            print_warning("Eine E-Mail wurde bereits innerhalb des definierten Zeitrahmens gesendet. Kein neuer Versand.")
            return
    else:
        last_email_time = None


    cutoff_time = datetime.now() - timedelta(hours=timeframe_hours)
    # Dateien im Zeitrahmen finden
    csv_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]

    # Dateien sortieren
    csv_files.sort(key=lambda f: os.path.getctime(os.path.join(import_dir, f)), reverse=True)

    # Datei suchen, die älter als der Zeitrahmen ist
    print_info(f"Suche neueste WebUntis-Import-Datei mit einem Mintestalter von {timeframe_hours} Stunden.")
    previous_file = next(
        (f for f in csv_files if datetime.fromtimestamp(os.path.getmtime(os.path.join(import_dir, f))) < cutoff_time),
        None
    )

    # Neueste Datei finden
    print_info(f"Suche neuste WebUntis-Import-Datei.")
    latest_file = csv_files[0] if csv_files else None

    if not latest_file:
        print_warning("Keine Dateien gefunden.")
        return

    if not previous_file:
        print_warning(f"Keine Dateien gefunden, die älter als {timeframe_hours} Stunden sind.")
        return

    print_info(f"  Vergleiche: '{previous_file}' → '{latest_file}'")

    previous_data = read_csv(os.path.join(import_dir, previous_file)) if previous_file else {}
    latest_data = read_csv(os.path.join(import_dir, latest_file))

    changes = []
    new_count = 0
    change_count = 0
    # 1. Änderungen und neue Schüler prüfen
    for student_id, latest_student in latest_data.items():
        previous_student = previous_data.get(student_id)
        
        if not previous_student:
            new_count += 1
            # Neuer Schüler - als Änderung markieren für die Historie
            changes.append({
                "student_id": student_id,
                "name": f"{latest_student.get('Vorname', '')} {latest_student.get('Nachname', '')}",
                "current_class": latest_student.get('Klasse', ''),
                "changes": {"__SYSTEM__": {"old": "N/A", "new": "Neu"}},
                "row": latest_student
            })
            continue

        row_changed = False
        updated_row = latest_student.copy()
        change_details = {}
        for field, new_value in latest_student.items():
            old_raw = previous_student.get(field, '').strip()
            new_raw = new_value.strip()

            if normalize_value(old_raw) != normalize_value(new_raw):
                row_changed = True
                change_count += 1
                updated_row[field] = f"{old_raw} -> {new_raw}"
                change_details[field] = {"old": old_raw, "new": new_raw}

        if row_changed:
            changes.append({
                "student_id": student_id,
                "name": f"{latest_student.get('Vorname', '')} {latest_student.get('Nachname', '')}",
                "current_class": latest_student.get('Klasse', ''),
                "changes": change_details,
                "row": updated_row
            })

    # 2. Fehlende Schüler prüfen
    missing_count = 0
    for student_id, previous_student in previous_data.items():
        if student_id not in latest_data:
            missing_count += 1
            changes.append({
                "student_id": student_id,
                "name": f"{previous_student.get('Vorname', '')} {previous_student.get('Nachname', '')}",
                "current_class": previous_student.get('Klasse', ''),
                "changes": {"__SYSTEM__": {"old": "Existierte", "new": "Fehlt"}},
                "row": previous_student
            })

    if not changes:
        print_warning("Keine Änderungen festgestellt.")
        return

    # Zusammenfassungs-Log
    print_success(f"Vergleich abgeschlossen: {new_count} neu, {missing_count} fehlen, {change_count} Änderungen.")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file_path = os.path.join(log_dir, f'ÄnderungsLog_{timestamp}.log')
    excel_file_path = os.path.join(xlsx_dir, f'ÄnderungsLog_{timestamp}.xlsx')

    # Änderungen in der Historien-Datenbank speichern
    history_manager.record_changes(changes, timestamp, previous_file, latest_file)

    # Log-Datei erstellen
    if not no_log:
        print_creation(f"Erstelle Text-Log-Datei: {log_file_path}")
        with open(log_file_path, 'w', encoding='utf-8') as log_file:
            for change in changes:
                log_file.write(f"Schüler: {change['name']} (ID: {change['student_id']})\n")
                for field, values in change['changes'].items():
                    log_file.write(f"  {field}: {values['old']} -> {values['new']}\n")
                log_file.write("\n")
        print_success(".log Log-Datei erfolgreich erstellt.")
    else:
        print_info("Log-Datei-Erstellung übersprungen (--no-log).")

    # Excel-Datei erstellen
    if no_xlsx:
        print_info("Excel-Datei-Erstellung übersprungen (--no-xlsx).")
        return
    print_creation(f"Erstelle Excel-Log-Datei: {excel_file_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Änderungen"

    headers = list(changes[0]["row"].keys())
    ws.append(headers)

    for change in changes:
        row = []
        for header in headers:
            value = change["row"].get(header, "")
            row.append(value)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if "->" in str(cell.value):
                cell.fill = red_fill

    wb.save(excel_file_path)
    print_success("Excel-Log-Datei erfolgreich erstellt.")

    # E-Mail mit den Änderungen senden
    print_info(f"Erstelle und sende E-Mails mit HTML-Tabelle und Excel-Datei im Anhang...")
    if changes:
        subject = "Änderungen im WebUntis Import"
        body = f"""
        <p>Die Dateien <strong>{previous_file}</strong> und <strong>{latest_file}</strong> wurden verglichen.</p>
        <p>Folgende Änderungen wurden festgestellt:</p>
        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; width: 100%;">
            <thead>
                <tr>
                    {"".join(f"<th>{header}</th>" for header in headers)}
                </tr>
            </thead>
            <tbody>
        """

        # Generiere die Tabelle mit den Änderungen
        for change in changes:
            body += "<tr>"
            for header in headers:
                value = change["row"].get(header, "")
                if "->" in value:  # Markiere Zellen mit Änderungen
                    body += f"<td style='background-color: #FFCCCC;'>{value}</td>"
                else:
                    body += f"<td>{value}</td>"
            body += "</tr>"

        body += """
            </tbody>
        </table>
        <p>Die vollständige Excel-Datei mit den Änderungen ist im Anhang enthalten.</p>
        """

        try:
            send_email(
                subject=subject,
                body=body,
                to_addresses=[admin_email],
                attachment_path=excel_file_path
            )
            print_success(f"E-Mail mit Änderungen an {admin_email} gesendet.")

            # Zeitstempel des E-Mail-Versands speichern
            print_info(f"Schreiben des Zeitstempels des Versands in {email_log_path} zur Verhinderung weiter E-Mails innerhalb der nächsten {timeframe_hours} Stunden...")
            with open(email_log_path, "w") as email_log:
                email_log.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            print_success(f"Zeit [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]notiert.")
        except Exception as e:
            print_error(f"Fehler beim Senden der E-Mail: {e}")


def read_classes(classes_dir, teachers_dir, return_teachers=False):
    # 0) WebUntis API Vorbereitung (Hybrid-Modus)
    config_api = configparser.ConfigParser()
    safe_read_config(config_api, 'email_settings.ini')
    use_api = config_api.getboolean('WebUntisAPI', 'use_api', fallback=False)
    
    teachers = {}
    classes_by_name = {}
    api_success = False

    if use_api:
        from webuntis_api import WebUntisClient
        try:
            print_info("Lese Daten über WebUntis API...")
            client = WebUntisClient(
                server_url=config_api.get('WebUntisAPI', 'server_url'),
                school=config_api.get('WebUntisAPI', 'school'),
                user=config_api.get('WebUntisAPI', 'user'),
                password=config_api.get('WebUntisAPI', 'password'),
                client_name=config_api.get('WebUntisAPI', 'client_name', fallback='Schild-WebUntis-Tool')
            )
            if client.authenticate():
                # 1) Lehrer laden
                api_teachers = client.get_teachers()
                for t in api_teachers:
                    t_name = t.get('name')
                    if t_name:
                        teachers[t_name] = {
                            'email': t.get('address.email') or t.get('email') or '',
                            'forename': t.get('foreName', ''),
                            'longname': t.get('longName', '')
                        }
                
                # 2) Klassen laden
                api_classes = client.get_classes()
                for c in api_classes:
                    c_name = (c.get('name') or "").strip().lower()
                    if c_name:
                        classes_by_name[c_name] = {
                            'Klassenlehrkraft_1': '', # API liefert oft keine KL direkt in getKlassen
                            'Klassenlehrkraft_1_Email': '',
                            'Klassenlehrkraft_2': '',
                            'Klassenlehrkraft_2_Email': '',
                            'api_data': True # Markierung für Admin-Warnings
                        }
                
                client.logout()
                api_success = True
                print_success(f"Daten erfolgreich über API geladen ({len(teachers)} Lehrer, {len(classes_by_name)} Klassen).")
                if return_teachers:
                    return classes_by_name, teachers
                return classes_by_name
        except Exception as e:
            print_warning(f"WebUntis API Fehler (Fallback auf CSV): {e}")

    if not api_success:
        # --- BESTEHENDER CSV-LOGIK-BLOCK ---
        # Funktion zum Einlesen der Klassen- und Lehrerdaten
        print_info("Lese Klassen- und Lehrerdaten aus CSV ein...")

        # Prüfen, ob der Ordner für Klassendaten existiert
        if not os.path.exists(classes_dir):
            print_warning(f"Warnung: Der Ordner '{classes_dir}' existiert nicht.")
            dummy_classes = {
                "dummy_class_1": {
                    "Klassenlehrkraft_1": "Max Muster",
                    "Klassenlehrkraft_1_Email": "max.muster@example.com",
                    "Klassenlehrkraft_2": "Erika Beispiel",
                    "Klassenlehrkraft_2_Email": "erika.beispiel@example.com",
                }
            }
            return (dummy_classes, {}) if return_teachers else dummy_classes

    # Neueste Klassen-CSV-Datei finden
    class_csv_files = [f for f in os.listdir(classes_dir) if f.lower().endswith('.csv')]
    if not class_csv_files:
        print_warning(f"Warnung: Keine Klassen-CSV-Dateien im Ordner '{classes_dir}' gefunden.")
        dummy_classes = {
            "dummy_class_1": {
                "Klassenlehrkraft_1": "Max Muster",
                "Klassenlehrkraft_1_Email": "max.muster@example.com",
                "Klassenlehrkraft_2": "Erika Beispiel",
                "Klassenlehrkraft_2_Email": "erika.beispiel@example.com",
            }
        }
        return (dummy_classes, {}) if return_teachers else dummy_classes

    newest_class_file = max(class_csv_files, key=lambda f: os.path.getctime(os.path.join(classes_dir, f)))
    print_info(f"Verwende Klassen-Datei: {newest_class_file}")

    # Prüfen, ob der Ordner für Lehrerdaten existiert
    if not os.path.exists(teachers_dir):
        print_warning(f"Warnung: Der Ordner '{teachers_dir}' existiert nicht.")
        teachers = {
            "dummy_teacher_1": {"email": "lehrer1@example.com", "forename": "Anna", "longname": "Lehrkraft"},
            "dummy_teacher_2": {"email": "lehrer2@example.com", "forename": "Tom", "longname": "Muster"},
        }
    else:
        # Neueste Lehrkräfte-CSV-Datei finden
        teacher_csv_files = [f for f in os.listdir(teachers_dir) if f.lower().endswith('.csv')]
        if not teacher_csv_files:
            print_warning(f"Warnung: Keine Lehrkräfte-CSV-Dateien im Ordner '{teachers_dir}' gefunden.")
            teachers = {
                "dummy_teacher_1": {"email": "lehrer1@example.com", "forename": "Anna", "longname": "Lehrkraft"},
                "dummy_teacher_2": {"email": "lehrer2@example.com", "forename": "Tom", "longname": "Muster"},
            }
        else:
            newest_teacher_file = max(teacher_csv_files, key=lambda f: os.path.getctime(os.path.join(teachers_dir, f)))
            print_info(f"Verwende Lehrkräfte-Datei: {newest_teacher_file}")

            # Lehrkräfte in ein Dictionary einlesen
            teachers = {}
            try:
                with open(os.path.join(teachers_dir, newest_teacher_file), 'r', newline='', encoding='utf-8-sig') as teacher_file:
                    teacher_reader = csv.DictReader(teacher_file, delimiter='\t')
                    for row in teacher_reader:
                        teacher_name = row['name']
                        teachers[teacher_name] = {
                            'email': row.get('address.email', ''),
                            'forename': row.get('foreName', ''),
                            'longname': row.get('longName', '')
                        }
            except Exception as e:
                print_error(f"Fehler beim Einlesen der Lehrkräfte-Datei: {e}")
                teachers = {}  # Fallback auf leere Lehrer-Daten

    # --- ENDE CSV-LOGIK-BLOCK ---

    # Klassen-CSV-Datei einlesen und Lehrkräfte-Emails zuordnen
    print_info("Verarbeite Klassen-CSV-Datei und ordne Lehrkräfte zu...")
    classes_by_name = {}
    with open(os.path.join(classes_dir, newest_class_file), 'r', newline='', encoding='utf-8-sig') as class_file:
        class_reader = csv.reader(class_file, delimiter=';')
        header = next(class_reader)

        # Die beiden Klassenlehrkraft-Spalten anhand ihrer Position (8. und 9. Spalte) identifizieren
        class_teacher_columns = [7, 8]  # 0-basierter Index: 8. und 9. Spalte

        for row in class_reader:
            emails = [teachers.get(row[idx] if len(row) > idx else '', 'Keine E-Mail gefunden') for idx in class_teacher_columns]
            row.extend(emails)

            kl1 = row[7] if len(row) > 7 else ''
            kl2 = row[8] if len(row) > 8 else ''
            kl_name = row[2].strip().lower() if len(row) > 2 else ''

            if not kl_name:
                continue

            # Klasseninformationen speichern
            classes_by_name[kl_name] = {
                'Klassenlehrkraft_1': f"{teachers.get(kl1, {}).get('forename', '')} {teachers.get(kl1, {}).get('longname', '')}",
                'Klassenlehrkraft_1_Email': teachers.get(kl1, {}).get('email', 'Keine E-Mail gefunden'),
                'Klassenlehrkraft_2': f"{teachers.get(kl2, {}).get('forename', '')} {teachers.get(kl2, {}).get('longname', '')}",
                'Klassenlehrkraft_2_Email': teachers.get(kl2, {}).get('email', 'Keine E-Mail gefunden')
            }
    print_success("Klassen- und Lehrerdaten erfolgreich eingelesen.")
    return (classes_by_name, teachers) if return_teachers else classes_by_name


def read_students(use_abschlussdatum=False):
    # Funktion zum Einlesen der Schülerdaten aus der neuesten CSV-Datei im aktuellen Verzeichnis
    print_info("Lese Schülerdaten ein...")
    import_dir = get_directory('import_directory', './WebUntis Importe')
    # Überprüfen, ob das Importverzeichnis existiert, falls nicht, erstellen
    if not os.path.exists(import_dir):
        os.makedirs(import_dir)
        print_success(f"Importverzeichnis '{import_dir}' wurde erstellt.")

    # Überprüfen, ob .csv Dateien im aktuellen Verzeichnis vorhanden sind
    schildexport_dir = get_directory('schildexport_directory', default='.')
    if schildexport_dir in ('.', '', None):
        schildexport_dir = os.getcwd()
    csv_files = [f for f in os.listdir(schildexport_dir) if f.lower().endswith('.csv')]

    if not csv_files:
        print_error("Keine CSV-Dateien im aktuellen Verzeichnis gefunden.")
        return [], {}

    # Neueste CSV-Datei bestimmen
    newest_file = max(csv_files, key=lambda f: os.path.getctime(os.path.join(schildexport_dir, f)))

    # Definierte Spalten, die gefiltert werden sollen
    columns_to_filter = ['Interne ID-Nummer', 'Nachname', 'Vorname', 'Geburtsdatum', 'Klasse', 'Klassenlehrer', 'Geschlecht', 'Entlassdatum', 'Aufnahmedatum', 'vorauss. Abschlussdatum', 'Schulpflicht erfüllt', 'Volljährig', 'E-Mail (privat)', 'Telefon-Nr.', 'Fax-Nr.', 'Straße', 'Postleitzahl', 'Ortsname']
    # Spalten für die Ausgabe
    output_columns = ['Interne ID-Nummer', 'Nachname', 'Vorname', 'Geburtsdatum', 'Klasse', 'Geschlecht', 'Entlassdatum', 'Aufnahmedatum', 'vorauss. Abschlussdatum', 'Schulpflicht', 'Volljährig', 'E-Mail (privat)', 'Telefon-Nr.', 'Fax-Nr.', 'Straße', 'Postleitzahl', 'Ortsname', 'Aktiv']

    output_data_students = [] # Liste für die Ausgabe der Schülerdaten
    students_by_id = {} # Dictionary für den schnellen Zugriff auf Schülerdaten per ID

    # Öffnen der neuesten CSV-Datei und Einlesen der Daten
    with open(os.path.join(schildexport_dir, newest_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        # Überprüfen, ob die notwendigen Spalten vorhanden sind
        header = [column for column in reader.fieldnames if column in columns_to_filter]
        header.append('Schulpflicht')
        header.append('Aktiv')
        output_data_students.append(output_columns)
        print_info("Beginne mit dem Verarbeiten der Schülerdaten...")
        line_num = 1 # Header
        for row in reader:
            line_num += 1
            filtered_row = {k: v for k, v in row.items() if k in columns_to_filter}

            # Schulpflicht-Logik (umkehren von 'Ja'/'Nein')
            if 'Schulpflicht erfüllt' in filtered_row:
                if filtered_row['Schulpflicht erfüllt'] == 'Ja':
                    filtered_row['Schulpflicht'] = 'Nein'
                elif filtered_row['Schulpflicht erfüllt'] == 'Nein':
                    filtered_row['Schulpflicht'] = 'Ja'

            # Datumskonvertierung für Entlassdatum
            entlassdatum = filtered_row.get('Entlassdatum', '')
            abschlussdatum = filtered_row.get('vorauss. Abschlussdatum', '')
            if entlassdatum and abschlussdatum:
                try:
                    entlassdatum_date = datetime.strptime(entlassdatum, '%d.%m.%Y')
                    abschlussdatum_date = datetime.strptime(abschlussdatum, '%d.%m.%Y')
                    if use_abschlussdatum and abschlussdatum_date < entlassdatum_date:
                        entlassdatum = abschlussdatum
                        print_info(f"Entlassdatum für Schüler {filtered_row['Nachname']}, {filtered_row['Vorname']} auf Abschlussdatum aktualisiert.")
                except ValueError:
                    print_error(f"Fehler (Zeile {line_num}) beim Konvertieren von Datumsangaben für Schüler {filtered_row['Nachname']}, {filtered_row['Vorname']}.")
                    pass # Ignoriere ungültige Datumsangaben
            elif not entlassdatum and abschlussdatum and use_abschlussdatum:
                entlassdatum = abschlussdatum
                print_creation(f"Entlassdatum (Zeile {line_num}) für Schüler {filtered_row['Nachname']}, {filtered_row['Vorname']} gesetzt auf Abschlussdatum.")

            filtered_row['Entlassdatum'] = entlassdatum
            
            # Schüler als aktiv markieren basierend auf dem Status
            filtered_row['Aktiv'] = 'Ja' if row['Status'] == '2' else 'Nein'

            # Klassenlehrer verarbeiten
            if 'Klassenlehrer' in reader.fieldnames:
                filtered_row['Klassenlehrer'] = row.get('Klassenlehrer', '').strip()
            else:
                filtered_row['Klassenlehrer'] = ''  # Standardwert, falls die Spalte fehlt

            # Füge die Daten zur Ausgabeliste hinzu
            output_data_students.append([filtered_row.get(col, '') for col in output_columns])
            # Speichere den Schüler im Dictionary nach Interner ID
            students_by_id[filtered_row['Interne ID-Nummer']] = filtered_row

    print_success("Schülerdaten erfolgreich eingelesen und verarbeitet.")
    return output_data_students, students_by_id

def create_warnings(classes_by_name, students_by_id):
    warnings = []
    import_dir = get_directory('import_directory', './WebUntis Importe')

    # Vorherige Importdatei finden
    output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    if output_files:
        newest_output_file = max(output_files, key=lambda f: os.path.getctime(os.path.join(import_dir, f)))
        with open(os.path.join(import_dir, newest_output_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                interne_id = row.get('Interne ID-Nummer')
                if interne_id in students_by_id:
                    # Vergleiche Entlassdatum
                    previous_entlassdatum = row.get('Entlassdatum')
                    new_entlassdatum = students_by_id[interne_id].get('Entlassdatum')
                    try:
                        previous_date = datetime.strptime(previous_entlassdatum, '%d.%m.%Y').date() if previous_entlassdatum else None
                        new_date = datetime.strptime(new_entlassdatum, '%d.%m.%Y').date() if new_entlassdatum else None
                        now_date = datetime.now().date()  # Importdatum ist aktuelles Datum

                        # Fall 1: Entlassdatum wird in die Zukunft verschoben und Importdatum liegt nach altem Entlassdatum
                        if previous_date and new_date and previous_date < new_date and now_date > previous_date:
                            if now_date < new_date:
                                undokumentierter_zeitraum_bis = now_date
                            else:
                                undokumentierter_zeitraum_bis = new_date

                            klasse = row['Klasse'].strip().lower()
                            klassen_info = classes_by_name.get(klasse, {})
                            if not klassen_info:
                                print_warning(f"Klasse '{klasse}' nicht in Klasseninformationen gefunden.")

                            # Warnung hinzufügen
                            warnings.append({
                                'Nachname': row['Nachname'],
                                'Vorname': row['Vorname'],
                                'Klasse': row.get('Klasse', 'N/A'),
                                'neues_entlassdatum': new_entlassdatum,
                                'altes_entlassdatum': previous_entlassdatum,
                                'Zeitraum_nicht_dokumentiert': f"{previous_entlassdatum} bis {undokumentierter_zeitraum_bis.strftime('%d.%m.%Y')}",
                                'Klassenlehrkraft_1': klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                                'Klassenlehrkraft_1_Email': klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                                'Klassenlehrkraft_2': klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                                'Klassenlehrkraft_2_Email': klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                                'Volljaehrig': students_by_id[interne_id].get('Volljährig', 'Unbekannt'),  # Volljaehrig Schlüssel hinzufügen
                                'Status': students_by_id[interne_id].get('Status_Text', 'Unbekannt'),  # Status Schlüssel hinzufügen
                                'warning_message': (
                                    "Das Entlassdatum wurde in die Zukunft verschoben, "
                                    "und es gibt einen nicht dokumentierten Zeitraum."
                                )
                            })
                    except ValueError:
                        print_error(f"Ungültiges Datum für Schüler {row['Nachname']}, {row['Vorname']}. Warnung nicht erstellt.")
    else:
        print_warning("Keine vorherige Importdatei zum Vergleich gefunden.")
    return warnings


def create_new_student_warnings(classes_by_name, current_students_by_id):
    warnings = []
    import_dir = get_directory('import_directory', './WebUntis Importe')
    
    # Alle CSV-Dateien aus dem Importverzeichnis ermitteln
    output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    
    # Wähle als Vergleichsdatei:
    if len(output_files) >= 1:
        # Sortiere absteigend nach Erstellungszeit und wähle die neueste Datei
        output_files.sort(key=lambda f: os.path.getctime(os.path.join(import_dir, f)), reverse=True)
        previous_file = output_files[0]
    else:
        print_warning("Keine Importdatei gefunden.")
        return warnings

    
    # Lese die Vergleichsdatei ein und sammle alle Schüler-IDs (normalisiert)
    previous_students = {}
    try:
        with open(os.path.join(import_dir, previous_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                student_id = row.get('Interne ID-Nummer')
                if student_id:
                    previous_students[student_id.strip()] = row
    except Exception as e:
        print_error(f"Fehler beim Einlesen der Vergleichsdatei: {e}")
        return warnings

    # Vergleiche die aktuellen Schülerdaten (students_by_id) mit denen der Vergleichsdatei
    for student_id, student in current_students_by_id.items():
        key = student_id.strip() if isinstance(student_id, str) else str(student_id)
        if key not in previous_students:
            klasse = student.get('Klasse', 'N/A').strip().lower()
            klassen_info = classes_by_name.get(klasse, {})
            warnings.append({
                'Nachname': student.get('Nachname', ''),
                'Vorname': student.get('Vorname', ''),
                'Klasse': student.get('Klasse', 'N/A'),
                'Aufnahmedatum': student.get('Aufnahmedatum', ''),
                'warning_message': "Neuer Schüler in den importierten Daten entdeckt.",
                'Klassenlehrkraft_1': klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                'Klassenlehrkraft_1_Email': klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                'Klassenlehrkraft_2': klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                'Klassenlehrkraft_2_Email': klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                'new_student': True
            })
    return warnings

def create_karteileichen_warnings(classes_by_name, current_students_by_id, admin_cache=None):
    warnings = []
    import_dir = get_directory('import_directory', './WebUntis Importe')
    
    # Alle CSV-Dateien aus dem Importverzeichnis ermitteln
    output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    
    # Wähle als Vergleichsdatei:
    if len(output_files) >= 1:
        # Sortiere absteigend nach Erstellungszeit und wähle die neueste Datei
        output_files.sort(key=lambda f: os.path.getctime(os.path.join(import_dir, f)), reverse=True)
        previous_file = output_files[0]
    else:
        print_warning("Keine vorherige Importdatei gefunden zum Vergleich für Karteileichen.")
        return warnings

    
    previous_students_names = []
    try:
        with open(os.path.join(import_dir, previous_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                student_id = row.get('Interne ID-Nummer', '').strip()
                if student_id and student_id not in current_students_by_id:
                    # Schüler ist in der alten Datei, aber nicht mehr in den aktuellen Imports
                    klasse = row.get('Klasse', 'N/A').strip().lower()
                    klassen_info = classes_by_name.get(klasse, {})
                    vorname = row.get('Vorname', '')
                    nachname = row.get('Nachname', '')
                    
                    warnings.append({
                        'Nachname': nachname,
                        'Vorname': vorname,
                        'Klasse': row.get('Klasse', 'N/A'),
                        'warning_message': "Schüler fehlt in aktueller Import-Datei gänzlich (Karteileiche).",
                        'Klassenlehrkraft_1': klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                        'Klassenlehrkraft_1_Email': klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                        'Klassenlehrkraft_2': klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                        'Klassenlehrkraft_2_Email': klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                        'karteileiche': True
                    })
                    previous_students_names.append(f"{vorname} {nachname} ({row.get('Klasse', 'N/A')})")
    except Exception as e:
        print_error(f"Fehler beim Einlesen der Vergleichsdatei für Karteileichen: {e}")
        return warnings

    if warnings and admin_cache is not None:
        names_str = ", ".join(previous_students_names)
        admin_cache.append({
            'Typ': 'Gelöschte Schüler (Karteileiche)',
            'Details': f"Folgende {len(warnings)} Schüler wurden in den neuen importierten Daten deklassiert / gelöscht: {names_str}. Verbleib prüfen."
        })

    return warnings




def create_class_change_warnings(classes_by_name, students_by_id, class_change_recipients="old"):
    # Funktion zum Erstellen von Warnungen bei Klassenwechseln
    warnings = []
    import_dir = get_directory('import_directory', './WebUntis Importe')

    # Vorherige Importdatei finden
    output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    if output_files:
        newest_output_file = max(output_files, key=lambda f: os.path.getctime(os.path.join(import_dir, f)))
        with open(os.path.join(import_dir, newest_output_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                interne_id = row.get('Interne ID-Nummer')
                if interne_id in students_by_id:
                    # Vergleiche Klasse
                    previous_class = row.get('Klasse')
                    new_class = students_by_id[interne_id].get('Klasse')
                    if previous_class and new_class and previous_class != new_class:
                        alte_klasse_str = previous_class.strip().lower()
                        neue_klasse_str = new_class.strip().lower()

                        alte_klassen_info = classes_by_name.get(alte_klasse_str, {})
                        neue_klassen_info = classes_by_name.get(neue_klasse_str, {})

                        if not alte_klassen_info:
                             print_warning(f"Alte Klasse '{previous_class}' nicht in Klasseninformationen gefunden.")
                        if not neue_klassen_info:
                             print_warning(f"Neue Klasse '{new_class}' nicht in Klasseninformationen gefunden.")

                        # Recipient-Logic bestimmen
                        recipients = []
                        if class_change_recipients in ["old", "both"]:
                             recipients.extend([alte_klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'), alte_klassen_info.get('Klassenlehrkraft_2_Email', 'N/A')])
                        if class_change_recipients in ["new", "both"]:
                             recipients.extend([neue_klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'), neue_klassen_info.get('Klassenlehrkraft_2_Email', 'N/A')])
                        
                        # E-Mails "N/A" und Duplikate filtern
                        recipients = list(set([r for r in recipients if r != 'N/A']))
                        if not recipients: recipients = ['N/A']

                        # Lehrkräfte-Texte für E-Mail vorbereiten
                        lehrkraefte_liste_alt = f"{alte_klassen_info.get('Klassenlehrkraft_1', 'N/A')} & {alte_klassen_info.get('Klassenlehrkraft_2', 'N/A')}"
                        lehrkraefte_liste_neu = f"{neue_klassen_info.get('Klassenlehrkraft_1', 'N/A')} & {neue_klassen_info.get('Klassenlehrkraft_2', 'N/A')}"
                        
                        # Ein Sammel-String für den E-Mail Body
                        lehrkraefte_tabelle = (
                            f"<strong>Beteiligte Lehrkräfte:</strong><br>"
                            f"Alte Klasse ({previous_class}): {lehrkraefte_liste_alt}<br>"
                            f"Neue Klasse ({new_class}): {lehrkraefte_liste_neu}"
                        )

                        # Warnung hinzufügen (Konsolidiert)
                        warnings.append({
                            'Nachname': row['Nachname'],
                            'Vorname': row['Vorname'],
                            'alte_klasse': previous_class,
                            'neue_klasse': new_class,
                            # Lehrer für die alte Klasse (für spezifische Platzhalter)
                            'alte_Klassenlehrkraft_1': alte_klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                            'alte_Klassenlehrkraft_1_Email': alte_klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                            'alte_Klassenlehrkraft_2': alte_klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                            'alte_Klassenlehrkraft_2_Email': alte_klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                            # Lehrer für die neue Klasse
                            'neue_Klassenlehrkraft_1': neue_klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                            'neue_Klassenlehrkraft_1_Email': neue_klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                            'neue_Klassenlehrkraft_2': neue_klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                            'neue_Klassenlehrkraft_2_Email': neue_klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                            # Standard-Lehrer Felder (Wiederhergestellt für Kompatibilität)
                            'Klassenlehrkraft_1': neue_klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                            'Klassenlehrkraft_1_Email': neue_klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                            'Klassenlehrkraft_2': neue_klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                            'Klassenlehrkraft_2_Email': neue_klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                            # Neue Felder für konsolidierte Mails
                            'anrede_global': "Kolleginnen und Kollegen" if class_change_recipients == "both" else neue_klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                            'lehrkraefte_tabelle': lehrkraefte_tabelle,
                            'recipients_list': recipients,
                            'Volljaehrig': students_by_id[interne_id].get('Volljährig', 'Unbekannt'),
                            'Status': students_by_id[interne_id].get('Status_Text', 'Unbekannt'),
                            'warning_message': "Klassenwechsel muss im Digitalen Klassenbuch manuell korrigiert werden, da die Änderung im System ab dem aktuellen Tag gilt."
                        })
    else:
        print_warning("Keine vorherige Importdatei zum Vergleich gefunden.")
    return warnings
    print_info(f"Anzahl der erstellten Klassenwechsel-Warnungen: {len(warnings)}")
    return warnings

def create_admission_date_warnings(classes_by_name, students_by_id):
    # Funktion zum Erstellen von Warnungen bei Änderungen des Aufnahmedatums
    warnings = []
    import_dir = get_directory('import_directory', './WebUntis Importe')

    # Vorherige Importdatei finden
    output_files = [f for f in os.listdir(import_dir) if f.lower().endswith('.csv') and 'Fehlende' not in f]
    if output_files:
        newest_output_file = max(output_files, key=lambda f: os.path.getctime(os.path.join(import_dir, f)))
        with open(os.path.join(import_dir, newest_output_file), 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                interne_id = row.get('Interne ID-Nummer')
                if interne_id in students_by_id:
                    # Prüfen, ob sich das Aufnahmedatum geändert hat
                    previous_admission_date = row.get('Aufnahmedatum')
                    new_admission_date = students_by_id[interne_id].get('Aufnahmedatum')
                    try:
                        # Konvertiere Datumsangaben zu `date`-Objekten
                        prev_date_obj = datetime.strptime(previous_admission_date, '%d.%m.%Y').date() if previous_admission_date else None
                        new_date_obj = datetime.strptime(new_admission_date, '%d.%m.%Y').date() if new_admission_date else None
                        now_date = datetime.now().date()  # Importdatum ist aktuelles Datum

                        # Fall 2: Aufnahmedatum wird in die Vergangenheit verschoben und Importdatum liegt nach dem neuen Aufnahmedatum
                        if prev_date_obj and new_date_obj and new_date_obj < prev_date_obj and now_date > new_date_obj:
                            if now_date < prev_date_obj:
                                undokumentierter_zeitraum_bis = now_date
                            else:
                                undokumentierter_zeitraum_bis = prev_date_obj
                            klasse = row.get('Klasse', 'N/A').strip().lower()
                            klassen_info = classes_by_name.get(klasse, {})

                            # Warnung hinzufügen
                            warnings.append({
                                'Nachname': row['Nachname'],
                                'Vorname': row['Vorname'],
                                'Klasse': row.get('Klasse', 'N/A'),
                                'neues_aufnahmedatum': new_admission_date,
                                'altes_aufnahmedatum': previous_admission_date,
                                'Zeitraum_nicht_dokumentiert': f"{new_admission_date} bis {undokumentierter_zeitraum_bis.strftime('%d.%m.%Y')}",
                                'Klassenlehrkraft_1': klassen_info.get('Klassenlehrkraft_1', 'N/A'),
                                'Klassenlehrkraft_1_Email': klassen_info.get('Klassenlehrkraft_1_Email', 'N/A'),
                                'Klassenlehrkraft_2': klassen_info.get('Klassenlehrkraft_2', 'N/A'),
                                'Klassenlehrkraft_2_Email': klassen_info.get('Klassenlehrkraft_2_Email', 'N/A'),
                                'Volljaehrig': students_by_id[interne_id].get('Volljährig', 'Unbekannt'),  # Volljaehrig hinzufügen
                                'Status': students_by_id[interne_id].get('Status_Text', 'Unbekannt'),  # Status Schlüssel hinzufügen
                                'warning_message': (
                                    "Das Aufnahmedatum wurde in die Vergangenheit verschoben, "
                                    "und es gibt einen nicht dokumentierten Zeitraum."
                                )
                            })
                    except ValueError:
                        print_error(f"Ungültiges Datum für Schüler {row['Nachname']}, {row['Vorname']}. Warnung nicht erstellt.")
    else:
        print_warning("Keine vorherige Importdatei zum Vergleich gefunden.")
    return warnings

def print_warnings(warnings):
    if not warnings:
        print_info("  Keine.")
        return

    table = Table(box=box.SIMPLE_HEAD, show_header=True, header_style="bold cyan",
                  show_edge=False, pad_edge=True)
    table.add_column("Name",      style="yellow",  no_wrap=True)
    table.add_column("Klasse",    style="cyan",     no_wrap=True)
    table.add_column("Änderung",  style="yellow")
    table.add_column("Lehrkraft", style="dim")

    for warning in warnings:
        nachname = warning.get('Nachname', '').strip()
        vorname  = warning.get('Vorname', '').strip()
        name = f"{nachname}, {vorname}" if nachname and vorname else nachname or vorname or "(Unbekannt)"

        if 'neues_entlassdatum' in warning:
            klasse = warning.get('Klasse', '') or ''
            detail = f"Entlassdatum {warning['altes_entlassdatum']} → {warning['neues_entlassdatum']}"
        elif 'neues_aufnahmedatum' in warning:
            klasse = warning.get('Klasse', '') or ''
            detail = f"Aufnahmedatum {warning['altes_aufnahmedatum']} → {warning['neues_aufnahmedatum']}"
        elif 'neue_klasse' in warning:
            klasse = f"{warning.get('alte_klasse', '?')} → {warning['neue_klasse']}"
            detail = "Klassenwechsel"
        else:
            klasse = warning.get('Klasse', '') or ''
            detail = warning.get('warning_message', 'Warnung')

        if klasse in ('N/A',):
            klasse = ''

        kl1  = warning.get('Klassenlehrkraft_1', '') or ''
        kl1m = warning.get('Klassenlehrkraft_1_Email', '') or ''
        kl2  = warning.get('Klassenlehrkraft_2', '') or ''
        kl2m = warning.get('Klassenlehrkraft_2_Email', '') or ''

        lehrkraft_parts = []
        if kl1 and kl1 != 'N/A':
            lehrkraft_parts.append(f"{kl1}\n({kl1m})" if kl1m else kl1)
        if kl2 and kl2 != 'N/A':
            lehrkraft_parts.append(f"{kl2}\n({kl2m})" if kl2m else kl2)
        lehrkraft = "\n".join(lehrkraft_parts) if lehrkraft_parts else "N/A"

        table.add_row(name, klasse, detail, lehrkraft)

    _console.print(table)


def save_files(output_data_students, warnings, create_second_file, admin_warnings_cache, disable_import_file_creation=False, disable_import_file_if_admin_warning=False, enable_attestpflicht_column=False, enable_nachteilsausgleich_column=False):
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    import_dir = get_directory('import_directory', './WebUntis Importe')
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # 1) Attestpflicht-Spalte
    if enable_attestpflicht_column:
        print_info("Füge Attestpflicht-Spalte hinzu...")
        attest_ids = read_attest_ids_from_latest_file()
        headers = output_data_students[0]
        if "Attestpflicht" not in headers:
            headers.append("Attestpflicht")
        id_index = headers.index("Interne ID-Nummer")

        for row_i in range(1, len(output_data_students)):
            row = output_data_students[row_i]
            student_id = row[id_index]
            row.append("Ja" if student_id in attest_ids else "Nein")

    # 2) Nachteilsausgleich-Spalte
    if enable_nachteilsausgleich_column:
        print_info("Füge Nachteilsausgleich-Spalte hinzu...")
        nachteil_ids = read_nachteilsausgleich_ids_from_latest_file()
        headers = output_data_students[0]
        if "Nachteilsausgleich" not in headers:
            headers.append("Nachteilsausgleich")

        id_index = headers.index("Interne ID-Nummer")

        for row_i in range(1, len(output_data_students)):
            row = output_data_students[row_i]
            student_id = row[id_index]
            row.append("true" if student_id in nachteil_ids else "false")

    output_file = os.path.join(import_dir, f'WebUntis Import {now}.csv')
    second_output_file = os.path.join(import_dir, f'WebUntis Import {now}_Fehlende Entlassdatumsangaben.csv')

    # Sicherstellen, dass das Importverzeichnis existiert
    os.makedirs(import_dir, exist_ok=True)

    # Hauptausgabedatei speichern    
    # 1) Prüfen, ob der Nutzer generell keine Datei möchte („Nur hier verarbeiten“)
    if disable_import_file_creation:
        print_warning("Keine Importdatei – Nur hier verarbeiten ('disable_import_file_creation' = True).")
        return
    # 2) Prüfen, ob bei Admin-Warnungen nichts erstellt werden soll
    if disable_import_file_if_admin_warning and admin_warnings_cache:
        print_warning("Keine Importdatei – 'disable_import_file_if_admin_warning' ist aktiv und Admin-Warnungen liegen vor.")
        return
    # Hauptausgabedatei speichern
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile, delimiter=';')
        writer.writerows(output_data_students)
    print_creation(f"WebUntis-Importdatei gespeichert: {output_file}")

    # Zweite Ausgabedatei speichern, falls gewünscht
    if create_second_file:
        print_info("Erstelle zweite Ausgabedatei für fehlende Entlassdatumsangaben...")
        second_output_columns = ['Interne ID-Nummer', 'Nachname', 'Vorname', 'Geburtsdatum', 'Klasse', 'Geschlecht', 'Entlassdatum']
        second_output_data = []

        # Suche die neueste Datei im Hauptverzeichnis
        schildexport_dir = get_directory('schildexport_directory', default='.')
        if schildexport_dir in ('.', '', None):
            schildexport_dir = os.getcwd()
        csv_files = [f for f in os.listdir(schildexport_dir) if f.lower().endswith('.csv')]

        if not csv_files:
            print_error("Fehler: Keine CSV-Dateien im Hauptverzeichnis gefunden.")
            return

        # Daten für zweite Datei extrahieren
        newest_file = max(
            [os.path.join(schildexport_dir, f) for f in csv_files], 
            key=os.path.getctime
        )
        print_info(f"Verwende Datei für zweite Ausgabe:\n {newest_file}")
        with open(newest_file, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                if row['Status'] != '2' and not row.get('Entlassdatum'):
                    filtered_row = {k: v for k, v in row.items() if k in second_output_columns}
                    second_output_data.append([filtered_row.get(col, '') for col in second_output_columns])
        if second_output_data:
            with open(second_output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                writer.writerow(second_output_columns)  # Header für die zweite Datei
                writer.writerows(second_output_data)
            print_creation(f"Zweite Ausgabedatei gespeichert:\n {second_output_file}")
        else:
            print_warning("Keine Daten für die zweite Ausgabedatei gefunden.")
    else:
        print_info("Keine zweite Ausgabedatei da die Erstellung der zweiten Ausgabedatei deaktiviert wurde.")

    
def read_attest_ids_from_latest_file():
    """
    Sucht im 'attest_file_directory' (aus settings.ini) nach der neuesten .csv-Datei,
    liest daraus alle 'Interne ID-Nummer' Einträge in ein Set und gibt dieses zurück.
    """
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    attest_dir = config.get('Directories', 'attest_file_directory', fallback='./AttestpflichtDaten')

    if not os.path.exists(attest_dir):
        print_warning(f"Attest-Verzeichnis '{attest_dir}' existiert nicht. Keine Attestdaten vorhanden.")
        return set()

    csv_files = [f for f in os.listdir(attest_dir) if f.lower().endswith('.csv')]
    if not csv_files:
        print_warning("Keine CSV-Dateien im Attest-Verzeichnis gefunden. Keine Attestdaten vorhanden.")
        return set()

    # Neueste CSV-Datei bestimmen
    csv_files.sort(key=lambda x: os.path.getctime(os.path.join(attest_dir, x)), reverse=True)
    latest_file = csv_files[0]
    full_path = os.path.join(attest_dir, latest_file)

    # IDs einlesen
    return read_attest_ids(full_path)
def read_attest_ids(attest_file_path):
    """
    Liest alle Interne ID-Nummern aus der gegebenen Attest-CSV-Datei und gibt sie als Set zurück.
    Beispielhafte Spalte: 'Interne ID-Nummer'
    """
    ids = set()
    print_info(f"Lese Attestpflicht-Datei: {attest_file_path}")
    try:
        with open(attest_file_path, 'r', encoding='utf-8-sig', newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                sid = row.get("Interne ID-Nummer", "").strip()
                if sid:
                    ids.add(sid)
    except Exception as e:
        print_error(f"Fehler beim Lesen der Attestpflicht-Datei '{attest_file_path}': {e}")
    print_info(f"Attestpflicht-Datei eingelesen – {len(ids)} betroffene Schüler-IDs.")
    return ids

def read_nachteilsausgleich_ids_from_latest_file():
    """
    Sucht im 'nachteilsausgleich_file_directory' aus settings.ini
    nach der neuesten .csv-Datei und liefert die IDs als Set zurück.
    """
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    nad_dir = config.get('Directories', 'nachteilsausgleich_file_directory', fallback='./NachteilsausgleichDaten')

    if not os.path.exists(nad_dir):
        print_warning(f"Nachteilsausgleich-Verzeichnis '{nad_dir}' existiert nicht. Alle = 'false'.")
        return set()

    csv_files = [f for f in os.listdir(nad_dir) if f.lower().endswith('.csv')]
    if not csv_files:
        print_warning("Keine CSV-Dateien im Nachteilsausgleich-Verzeichnis gefunden. Alle = 'false'.")
        return set()

    # Neueste Datei
    csv_files.sort(key=lambda x: os.path.getctime(os.path.join(nad_dir, x)), reverse=True)
    latest_file = csv_files[0]
    full_path = os.path.join(nad_dir, latest_file)

    return read_nachteilsausgleich_ids(full_path)

def read_nachteilsausgleich_ids(file_path):
    """
    Liest 'Interne ID-Nummer' aus der CSV und gibt sie als Set zurück.
    """
    ids = set()
    print_info(f"Lese Nachteilsausgleich-Datei: {file_path}")
    try:
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                sid = (row.get("Interne ID-Nummer") or "").strip()
                if sid:
                    ids.add(sid)
    except Exception as e:
        print_error(f"Fehler beim Lesen Nachteilsausgleich-Datei: {e}")
    print_info(f"Nachteilsausgleich-Datei eingelesen. Anzahl IDs: {len(ids)}")
    return ids

from collections import defaultdict

def create_class_sizes_file(students_by_id):
    """
    Erzeugt eine CSV mit den Spalten:
      Klasse, Schüler (männlich), Schüler (weiblich), Schüler (divers), Schüler (gesamt)
    und berücksichtigt nur Schüler mit Status == '2' (also Aktiv == 'Ja').
    Die Datei wird in das class_size_directory aus den settings.ini abgelegt.
    """
    print_info("Erstelle Klassengrößen-Auswertung...")

    # 1) Lese aus settings.ini das Verzeichnis
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    class_size_dir = config.get('Directories', 'class_size_directory', fallback="./ClassSizes")

    # Stelle sicher, dass der Ordner existiert
    os.makedirs(class_size_dir, exist_ok=True)

    # 2) Aggregation: { klasse_name: {"m": 0, "w": 0, "d": 0} }
    class_counts = defaultdict(lambda: {"m": 0, "w": 0, "d": 0})

    for student_id, student_data in students_by_id.items():
        # Nur Schüler mit Status=2 => 'Aktiv' == 'Ja'
        if student_data.get("Aktiv") == "Ja":
            klasse = (student_data.get("Klasse") or "").strip()
            geschlecht = (student_data.get("Geschlecht") or "").lower()  # 'm', 'w', 'd'
            if geschlecht in ("m", "w", "d"):
                class_counts[klasse][geschlecht] += 1

    # 3) CSV-Datei schreiben
    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file_path = os.path.join(class_size_dir, f"Klassengroessen_{now_str}.csv")
    headers = ["Klasse", "Schüler (männlich)", "Schüler (weiblich)", "Schüler (divers)", "Schüler (gesamt)"]

    try:
        with open(output_file_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)

            # Sortiere Klassen wahlweise alphabetisch
            for klasse in sorted(class_counts.keys()):
                counts = class_counts[klasse]
                male = counts["m"]
                female = counts["w"]
                diverse = counts["d"]
                total = male + female + diverse
                writer.writerow([klasse, male, female, diverse, total])
        print_success(f"Klassengrößen-Datei erstellt: {output_file_path}")
    except Exception as e:
        print_error(f"Fehler beim Erstellen der Klassengrößen-Datei: {e}")







