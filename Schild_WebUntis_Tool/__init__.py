import os  # Betriebssystemfunktionen wie Pfadoperationen
import sys  # Systemfunktionen und -parameter

# Basisverzeichnis explizit zum Suchpfad hinzufügen, um relative Importe zu ermöglichen
base_dir = os.path.abspath(os.path.dirname(__file__))
sys.path.append(base_dir)

# Zwingend UTF-8 für die Standardausgabe konfigurieren, um Abstürze durch Emojis (wie ❌, ✅, ℹ️) auf veralteten Windows Konsolen in der .exe zu verhindern.
try:
    if sys.stdout and hasattr(sys.stdout, 'encoding') and sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

import csv  # Lesen und Schreiben von CSV-Dateien
import configparser  # Verarbeiten von Konfigurationsdateien im INI-Format
import webbrowser  # Öffnen von Webbrowsern
import threading  # Multithreading-Funktionen
import argparse  # Parsen von Kommandozeilenargumenten
import secrets
import traceback
import logging

# Logger konfigurieren
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
import history_manager

from utils import safe_read_config

import winshell  # Interaktion mit der Windows-Shell (z.B. Erstellen von Verknüpfungen)
import pythoncom  # Python COM-Schnittstelle für Windows
import threading  # Multithreading-Funktionen
import werkzeug  # Werkzeug-Bibliothek für WSGI-Anwendungen (von Flask verwendet)
import tkinter as tk  # GUI-Toolkit für die Dateiauswahl-Dialoge
import colorama  # Ausgabe von farbigem Text in der Konsole
from datetime import datetime  # Arbeiten mit Datum und Uhrzeit
from flask import Flask, render_template, request, jsonify, session, send_from_directory  # Flask-Webframework
from main import run, read_students, read_classes, compare_timeframe_imports, validate_imports  # Funktionen aus eigenen Modulen importieren
from smtp import send_email  # Funktion zum Versenden von E-Mails aus eigenem Modul
from waitress import serve  # WSGI-Server zum Bereitstellen der Flask-Anwendung
from tkinter import filedialog  # Datei- und Verzeichnisauswahl-Dialoge
from werkzeug.utils import secure_filename  # Sichere Dateinamen-Verarbeitung
from colorama import Fore, Back, Style, init  # Farb- und Stildefinitionen für die Konsolenausgabe

# Colorama initialisieren
init(autoreset=True)

# Thread-sicherer Zugriff
console_lock = threading.Lock()

def thread_safe_print(color, message):
    with console_lock:
        print(f"{color}{message}{Style.RESET_ALL}", flush=True)

# Hilfsfunktionen
def print_error(message):
    thread_safe_print(Fore.RED, f"❌ {message}")

def print_warning(message):
    thread_safe_print(Fore.YELLOW, f"⚠️ {message}")

def print_admin_warning(message):
    thread_safe_print(Fore.LIGHTRED_EX, f"❗ {message}")

def print_warningtext(message):
    thread_safe_print(Fore.MAGENTA, f"👾 {message}")

def print_success(message):
    thread_safe_print(Fore.GREEN, f"✅ {message}")

def print_info(message):
    thread_safe_print(Fore.CYAN, f"ℹ️ {message}")

def print_creation(message):
    thread_safe_print(Fore.WHITE, f"✨ {message}")

def print_banner():
    with console_lock:
        print(f"{Fore.CYAN}{'='*60}")
        print(f"{Fore.CYAN}🚀 {Fore.WHITE}{Style.BRIGHT}Schild-WebUntis-Tool v3.0 {Fore.CYAN}aktiviert")
        print(f"{Fore.CYAN}{'='*60}")
        print(f"{Fore.CYAN}📂 Verzeichnisse:")
        print(f"   ℹ️  Import:  {get_directory('import_directory', './WebUntis Importe')}")
        print(f"   ℹ️  Logs:    {get_directory('log_directory', './Logs')}")
        print(f"   ℹ️  Excel:   {get_directory('xlsx_directory', './ExcelExports')}")
        print(f"{Fore.CYAN}{'='*60}{Style.RESET_ALL}\n", flush=True)

def get_directory(key, default=None):
    # Hilfsfunktion zum Abrufen von Verzeichnispfaden aus der Konfigurationsdatei
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    return config.get('Directories', key, fallback=default)



def resource_path(relative_path):
    # Gibt den Pfad zu einer Ressource zurück, funktioniert für dev und bei PyInstaller
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Flask-App initialisieren
app = Flask(__name__,
            template_folder=os.path.join(base_dir, 'templates'),
            static_folder=os.path.join(base_dir, 'static'))
app.secret_key = secrets.token_hex(24) # Generierung eines zufälligen Secret Keys für die Session

# Globale Variablen für Warnungen und generierte E-Mails
global warnings_cache, generated_emails_cache, admin_warnings_cache

# Globale Caches für Warnungen und generierte E-Mails
warnings_cache = []  # Cache für Warnungen
generated_emails_cache = []  # Cache für generierte E-Mails
admin_warnings_cache = []

# Start der Datenverarbeitung über das Kommandozeilen-Argument --process (nicht WebEnd-Button) heraus.
def process_data(no_log=False, no_xlsx=False):
    # Konfigurationsdatei einlesen
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    
    # Werte aus der Konfigurationsdatei laden, da in diesem Fall das WebEnd nicht immer geöffnet ist.
    use_abschlussdatum = config.getboolean('ProcessingOptions', 'use_abschlussdatum', fallback=False)
    create_second_file = config.getboolean('ProcessingOptions', 'create_second_file', fallback=False)
    create_class_size_file = config.getboolean('ProcessingOptions', 'create_class_size_file', fallback=True)
    enable_attestpflicht_column = config.getboolean('ProcessingOptions', 'enable_attestpflicht_column', fallback=False)
    enable_nachteilsausgleich_column = config.getboolean('ProcessingOptions', 'enable_nachteilsausgleich_column', fallback=False)
    disable_import_file_creation= config.getboolean('ProcessingOptions', 'disable_import_file_creation', fallback=False)
    disable_import_file_if_admin_warning= config.getboolean('ProcessingOptions', 'disable_import_file_if_admin_warning', fallback=False)
    warn_entlassdatum = config.getboolean('ProcessingOptions', 'warn_entlassdatum', fallback=True)
    warn_aufnahmedatum = config.getboolean('ProcessingOptions', 'warn_aufnahmedatum', fallback=True)
    warn_klassenwechsel = config.getboolean('ProcessingOptions', 'warn_klassenwechsel', fallback=True)
    warn_new_students = config.getboolean('ProcessingOptions', 'warn_new_students', fallback=True)
    warn_karteileichen = config.getboolean('ProcessingOptions', 'warn_karteileichen', fallback=False)
    
    # CLI Parameter für class_change_recipients hat Vorrang vor settings.ini, falls gesetzt
    global cli_args
    if cli_args and cli_args.get("class_change_recipients"):
        class_change_recipients = cli_args.get("class_change_recipients")
    else:
        class_change_recipients = config.get('ProcessingOptions', 'class_change_recipients', fallback='old')
    
    # Aktuelle Einstellungen in der Konsole ausgeben
    print_info("Beginne Verarbeitung mit folgenden Optionen:...")
    print_info(f"use_abschlussdatum: {use_abschlussdatum}")
    print_info(f"create_second_file: {create_second_file}")
    print_info(f"create_class_size_file: {create_class_size_file}")
    print_info(f"enable_attestpflicht_column: { enable_attestpflicht_column}")
    print_info(f"warn_entlassdatum: {warn_entlassdatum}")
    print_info(f"warn_aufnahmedatum: {warn_aufnahmedatum}")
    print_info(f"warn_klassenwechsel: {warn_klassenwechsel}")
    print_info(f"class_change_recipients: {class_change_recipients}")
    print_info(f"warn_new_students: {warn_new_students}")
    print_info(f"warn_karteileichen: {warn_karteileichen}")

    # Datenverarbeitung starten
    all_warnings = run(                         #Hier wird die def_run aus der main.py mit den erfassten Einstellungen abgerufen und ausgeführt.
        use_abschlussdatum=use_abschlussdatum,
        create_second_file=create_second_file,
        enable_attestpflicht_column=enable_attestpflicht_column,
        create_class_size_file= create_class_size_file,
        disable_import_file_creation=disable_import_file_creation,
        disable_import_file_if_admin_warning=disable_import_file_if_admin_warning,
        warn_entlassdatum=warn_entlassdatum,
        warn_aufnahmedatum=warn_aufnahmedatum,
        warn_klassenwechsel=warn_klassenwechsel,
        warn_new_students=warn_new_students,
        warn_karteileichen=warn_karteileichen,
        class_change_recipients=class_change_recipients,
        no_log=no_log,
        no_xlsx=no_xlsx,
        admin_warnings_cache=admin_warnings_cache
    )
    print_success("Verarbeitung über die Konsole abgeschlossen.")
    return all_warnings

# Standard-Vorlagen für E-Mails
DEFAULT_TEMPLATES = {
    "entlassdatum": {
        "subject": "Webuntis-Hinweis: Entlassdatum-Problem bei $Vorname $Nachname",
        "body": "<p>Sehr geehrter/Sehr geehrte Herr/Frau $Klassenlehrkraft_1,</p><p>Es gibt ein Problem mit dem Entlassdatum des Schülers/der Schülerin <strong>$Vorname $Nachname</strong> aus der Klasse <strong>$Klasse</strong>.</p><p></p><p><strong>Neues Entlassdatum:</strong> $neues_entlassdatum</p><p><strong>Altes Entlassdatum:</strong> $altes_entlassdatum</p><p></p><p>$zeitraum_text</p><p></p><p>In dieser Zeit war der Schüler/die Schülerin nun offiziel teil der Klasse, jedoch nicht im digitalen Klassenbuch dokumentierbar. Dies muss nun nachgeholt werden.</p><p></p><p><strong>Klassenlehrkraft 1:</strong> $Klassenlehrkraft_1, E-Mail: $Klassenlehrkraft_1_Email</p><p><strong>Klassenlehrkraft 2:</strong> $Klassenlehrkraft_2, E-Mail: $Klassenlehrkraft_2_Email</p><p></p><p><strong>Hinweis:</strong> Es ist nicht möglich auf diese E-Mail Adresse zu antworten.</p><p></p><p>Mit freundlichen Grüßen,</p><p>Das WebUntis Team</p>"
    },
    "aufnahmedatum": {
        "subject": "Webuntis-Hinweis: Aufnahmedatum-Problem bei $Vorname $Nachname",
        "body": "<p>Sehr geehrter/Sehr geehrte Herr/Frau $Klassenlehrkraft_1,</p><p>Das Aufnahmedatum des Schülers/der Schülerin <strong>$Vorname $Nachname</strong> aus der Klasse <strong>$Klasse</strong> hat sich geändert.</p><p></p><p><strong>Neues Aufnahmedatum:</strong> $neues_aufnahmedatum</p><p><strong>Altes Aufnahmedatum:</strong> $altes_aufnahmedatum</p><p></p><p>$zeitraum_text</p><p></p><p>In dieser Zeit war der Schüler/die Schülerin nun offiziel teil der Klasse, jedoch nicht im digitalen Klassenbuch dokumentierbar. Dies muss nun nachgeholt werden.</p><p></p><p><strong>Klassenlehrkraft 1:</strong> $Klassenlehrkraft_1, E-Mail: $Klassenlehrkraft_1_Email</p><p><strong>Klassenlehrkraft 2:</strong> $Klassenlehrkraft_2, E-Mail: $Klassenlehrkraft_2_Email</p><p></p><p><strong>Hinweis:</strong> Es ist nicht möglich auf diese E-Mail Adresse zu antworten.</p><p></p><p>Mit freundlichen Grüßen,</p><p>Das WebUntis Team</p>"
    },
    "klassenwechsel": {
        "subject": "Webuntis-Hinweis: Klassenwechsel bei $Vorname $Nachname",
        "body": "<p>Sehr geehrte/r $anrede_global,</p><p>Es gab einen Klassenwechsel des Schülers/der Schülerin <strong>$Vorname $Nachname</strong>.</p><p>Sofern dieser Klassenwechsel nicht am heutigen Tag stattfand, informieren Sie bitte das WebUntis-Team über die Notwendigkeit einer Korrektur. </p><p>Liegt der Wechsel in der Vergangenheit müssen anschließend die Tage zwischen heute und diesem Wechsel nachdokumentiert werden.</p><p></p><p>$lehrkraefte_tabelle</p><p></p><p><strong>Hinweis:</strong> Es ist nicht möglich auf diese E-Mail Adresse zu antworten.</p><p></p><p>Mit freundlichen Grüßen,</p><p>Das WebUntis Team</p>"
    },
    "new_student": {
        "subject": "Webuntis-Hinweis: Neuer Schüler $Vorname $Nachname",
        "body": "<p>Sehr geehrte/r $Klassenlehrkraft_1,</p><p>Der Schüler/die Schülerin <strong>$Vorname $Nachname</strong> aus der Klasse <strong>$Klasse</strong> wurde als neu in den importierten Daten erkannt.</p><p>Bitte überprüfen Sie die Daten im digitalen Klassenbuch.</p><p>Mit freundlichen Grüßen,</p><p>Das WebUntis Team</p>"
    },
    "karteileiche": {
        "subject": "Webuntis-Hinweis: Schüler fehlt/gelöscht $Vorname $Nachname",
        "body": "<p>Sehr geehrte/r $Klassenlehrkraft_1,</p><p>Der Schüler/die Schülerin <strong>$Vorname $Nachname</strong> (Klasse <strong>$Klasse</strong>) taucht in den von SchILD importierten Daten (aktives Schuljahr) nicht mehr auf.</p><p>Dies wird dazu führen, dass seine Daten, darunter auch sein Entlassdatum, sein Status (inkl. von Aktiv nach Abschluss/Abgang) nicht mehr aktualisiert werden. Sofern das in Ordnung ist, ist keine weitere Aktion erforderlich. Falls nicht, prüfen Sie bitte den Verbleib des Schülers.</p><p>Mit freundlichen Grüßen,</p><p>Das WebUntis Team</p>"
    }
}

# Prüfen, ob die Konfigurationsdateien existieren, und sie bei Bedarf mit Standardwerten erstellen. Die Funktion wird unmittlbar bei Start des Servers unter "if __name__ == "__main__":" abgerufen.
def ensure_ini_files_exist():
    # Standardverzeichnisse definieren
    default_classes_dir = "Klassendaten"
    default_teachers_dir = "Lehrerdaten"
    default_log_dir = "Logs"
    default_xlsx_dir = "ExcelExports"
    default_import_dir = "WebUntis Importe"
    default_schildexport_dir = "."
    default_class_size_dir = "ClassSizes"
    default_attest_file_directory ="AttestpflichtDaten"
    default_nachteilsausgleich_file_directory ="NachteilsausgleichDaten"

    # Standard-Inhalt für settings.ini vorbereiten
    settings_ini_content = f"""[Directories]
classes_directory = {default_classes_dir}
teachers_directory = {default_teachers_dir}
log_directory = {default_log_dir}
xlsx_directory = {default_xlsx_dir}
import_directory = {default_import_dir}
schildexport_directory = {default_schildexport_dir}
class_size_directory = {default_class_size_dir}
attest_file_directory = {default_attest_file_directory}
nachteilsausgleich_file_directory = {default_nachteilsausgleich_file_directory}

[ProcessingOptions]
use_abschlussdatum = False
create_second_file = False
warn_entlassdatum = True
warn_aufnahmedatum = True
warn_klassenwechsel = True
warn_new_students = True
warn_karteileichen = False
create_class_size_file = False
timeframe_hours = 24
enable_attestpflicht_column = False
enable_nachteilsausgleich_column = False
disable_import_file_creation = False
disable_import_file_if_admin_warning = False

[mail]
# Empfänger nur bei Klassenwechsel-Warnungen
# gültig: old | new | both
class_change_recipients = both
"""

    # Standard-Inhalt für email_settings.ini vorbereiten
    email_settings_ini_content = f"""# Einstellungen für den E-Mail-Versand
# Passen Sie diese Einstellungen an Ihren SMTP-Server an.
# Beispiele:
# smtp_server = smtp.gmail.com # SMTP-Server-Adresse
# smtp_port = 587 # SMTP-Port (z. B. 587 für STARTTLS)
# smtp_user = ihrbenutzer@gmail.com # Benutzername für SMTP
# smtp_password = ihrpasswort # Passwort für SMTP
# smtp_encryption = starttls # Verschlüsselungsmethode (starttls oder ssl)
# admin_email = admin@example.com #E-Mail Adresse des Admins zum Erhalt spezieller Admin-Warnungen bei Nutzung über die Kommandozeile

[Email]
smtp_server = smtp.example.com  
smtp_port = 587  
smtp_user = user@example.com  
smtp_password = password  
smtp_encryption = starttls
admin_email = admin@example.com

# OAuth-Einstellungen (falls verwendet)
[OAuth]
use_oauth = False
credentials_path = ./config/credentials.json

# Vorlagen für generierte E-Mails
# Verwenden Sie Platzhalter wie {{Vorname}}, {{Nachname}}, {{Klasse}}, {{neues_entlassdatum}}, etc.

[Templates]
subject_entlassdatum = {DEFAULT_TEMPLATES['entlassdatum']['subject']}
body_entlassdatum = {DEFAULT_TEMPLATES['entlassdatum']['body']}
subject_aufnahmedatum = {DEFAULT_TEMPLATES['aufnahmedatum']['subject']}
body_aufnahmedatum = {DEFAULT_TEMPLATES['aufnahmedatum']['body']}
subject_klassenwechsel = {DEFAULT_TEMPLATES['klassenwechsel']['subject']}
body_klassenwechsel = {DEFAULT_TEMPLATES['klassenwechsel']['body']}
subject_new_student = {DEFAULT_TEMPLATES['new_student']['subject']}
body_new_student = {DEFAULT_TEMPLATES['new_student']['body']}
subject_karteileiche = {DEFAULT_TEMPLATES['karteileiche']['subject']}
body_karteileiche = {DEFAULT_TEMPLATES['karteileiche']['body']}
[WebUntisAPI]
use_api = False
server_url = https://neptun.webuntis.com/WebUntis/jsonrpc.do
school = schoolname
user = user
password = password
client_name = Schild-WebUntis-Tool
"""

    # Prüfen, ob settings.ini existiert, und ggf. erstellen
    settings_ini_exists = os.path.exists("settings.ini")
    if not settings_ini_exists:
        with open("settings.ini", "w", encoding="utf-8-sig") as file:
            file.write(settings_ini_content)
        print_success("Standard-Konfigurationsdatei 'settings.ini' wurde erstellt.")
    else:
        # Bestehende Konfiguration patchen (neue Felder hinzufügen)
        config = configparser.ConfigParser()
        if safe_read_config(config, "settings.ini"):
            updated = False
            # ProcessingOptions
            if not config.has_option('ProcessingOptions', 'warn_karteileichen'):
                config.set('ProcessingOptions', 'warn_karteileichen', 'False')
                updated = True
            
            if updated:
                with open("settings.ini", "w", encoding="utf-8-sig") as configfile:
                    config.write(configfile)
                print_info("settings.ini wurde um fehlende Einträge aktualisiert (Auto-Patcher).")

    # Prüfen, ob email_settings.ini existiert, und ggf. erstellen
    email_settings_ini_exists = os.path.exists("email_settings.ini")
    if not email_settings_ini_exists:
        with open("email_settings.ini", "w", encoding="utf-8-sig") as file:
            file.write(email_settings_ini_content)
        print_success("Standard-Konfigurationsdatei 'email_settings.ini' wurde erstellt.")
    else:
        # Bestehende Konfiguration patchen (neue Felder hinzufügen)
        config = configparser.ConfigParser()
        if safe_read_config(config, "email_settings.ini"):
            updated = False
            # Templates
            if not config.has_option('Templates', 'subject_karteileiche'):
                config.set('Templates', 'subject_karteileiche', 'Webuntis-Hinweis: Schüler fehlt/gelöscht $Vorname $Nachname')
                updated = True
            if not config.has_option('Templates', 'body_karteileiche'):
                config.set('Templates', 'body_karteileiche', DEFAULT_TEMPLATES['karteileiche']['body'])
                updated = True
            
            # Ensure other templates are also present (Silent Update)
            for t_type in ['entlassdatum', 'aufnahmedatum', 'klassenwechsel', 'new_student']:
                if not config.has_option('Templates', f'subject_{t_type}'):
                    config.set('Templates', f'subject_{t_type}', DEFAULT_TEMPLATES[t_type]['subject'])
                    updated = True
                if not config.has_option('Templates', f'body_{t_type}'):
                    config.set('Templates', f'body_{t_type}', DEFAULT_TEMPLATES[t_type]['body'])
                    updated = True

            # Spezielles Update für Klassenwechsel (Force Update auf neues Tabellen-Format)
            if config.has_option('Templates', 'body_klassenwechsel'):
                current_body = config.get('Templates', 'body_klassenwechsel')
                if "$lehrkraefte_tabelle" not in current_body:
                    config.set('Templates', 'body_klassenwechsel', DEFAULT_TEMPLATES['klassenwechsel']['body'])
                    updated = True
            
            if not config.has_section('WebUntisAPI'):
                config.add_section('WebUntisAPI')
                config.set('WebUntisAPI', 'use_api', 'False')
                config.set('WebUntisAPI', 'server_url', 'https://neptun.webuntis.com/WebUntis/jsonrpc.do')
                config.set('WebUntisAPI', 'school', 'schoolname')
                config.set('WebUntisAPI', 'user', 'user')
                config.set('WebUntisAPI', 'password', 'password')
                config.set('WebUntisAPI', 'client_name', 'Schild-WebUntis-Tool')
                updated = True
            
            if updated:
                with open("email_settings.ini", "w", encoding="utf-8-sig") as configfile:
                    config.write(configfile)
                print_info("email_settings.ini wurde um fehlende Einträge aktualisiert (Auto-Patcher).")

    # Sicherstellen, dass die in settings.ini definierten Ordner existieren
    config = configparser.ConfigParser()
    if settings_ini_exists:
        safe_read_config(config, "settings.ini")

    directories = {
        "classes_directory": default_classes_dir,
        "teachers_directory": default_teachers_dir,
        "log_directory": default_log_dir,
        "xlsx_directory": default_xlsx_dir,
        "schildexport_directory": default_schildexport_dir,
        "class_size_directory": default_class_size_dir,
        "import_directory": default_import_dir,
        "attest_file_directory": default_attest_file_directory,
        "nachteilsausgleich_file_directory": default_nachteilsausgleich_file_directory,
    }

    for key, default_path in directories.items():
        directory = config.get("Directories", key, fallback=default_path, raw=True)
        directory = os.path.normpath(directory)
        print_info(f"Verwende Verzeichnis für '{key}': {directory}")
        if not os.path.isabs(directory):
            directory = os.path.abspath(directory)
            print_info(f"Relativer Pfad erkannt. Absoluter Pfad: {directory}")
        if not os.path.exists(directory):
            os.makedirs(directory)
            print_creation(f"Ordner '{directory}' wurde erstellt.")


    global admin_warnings_cache
    admin_warnings_cache = []

# Generieren der Admin-Warnungen bei für den Fall inkonsistenter Daten. Wird ebenfalls direkt beim Start des Servers unter "if __name__ == "__main__":" abegerufen, sofern kein --skip-admin-warnings verwendet wurde.
def admin_warnings(send_email_flag=False):

    # Admin-Warnungen werden erstellt
    print_info("Erstelle Admin-Warnungen...")

    # Haupt-Import-Datei einlesen
    students_output, students_by_id = read_students(use_abschlussdatum=False)

    # Konfigurationsdatei einlesen
    config = configparser.ConfigParser()
    safe_read_config(config, 'settings.ini')
    classes_dir = config.get('Directories', 'classes_directory')
    teachers_dir = config.get('Directories', 'teachers_directory')

    # Klassen- und Lehrkräfte-Daten einlesen
    classes_by_name, teachers = read_classes(classes_dir, teachers_dir, return_teachers=True)

    # Lehrer-Daten aus der Lehrerdatei extrahieren (Spalte "name")
    teacher_names = set(teachers.keys())

    # Überprüfung auf fehlende Klassen in der Klassen-Datei
    for student in students_by_id.values():
        klasse = student.get('Klasse', '').strip().lower()
        if klasse not in classes_by_name:
            admin_warnings_cache.append({
                'Typ': 'Fehlende Klasse in der Klassen-Datei',
                'Details': f"Die Klasse '{klasse}' aus der Haupt-Import-Datei existiert nicht in der Klassen-Datei.",
                'Schüler': f"{student.get('Vorname', '')} {student.get('Nachname', '')}"
            })

    # Überprüfung auf fehlende Klassenlehrkräfte in der Lehrkräfte-Datei
    for student in students_by_id.values():
        klassenlehrer = student.get('Klassenlehrer', '').strip()  # Klassenlehrer aus Haupt-Import        
        if klassenlehrer:  # Nur prüfen, wenn der Wert vorhanden ist
            if klassenlehrer not in teacher_names:  # Vergleich mit Lehrkräfte-Datei
                admin_warnings_cache.append({
                    'Typ': 'Fehlender Klassenlehrer in Lehrerdatei',
                    'Details': f"Der Klassenlehrer '{klassenlehrer}' aus der Haupt-Import-Datei existiert nicht in der Lehrkräfte-Datei.",
                    'Schüler': f"{student.get('Vorname', '')} {student.get('Nachname', '')}"
                })

    # Überprüfen, ob CSV-Dateien im Klassenverzeichnis vorhanden sind
    class_csv_files = [f for f in os.listdir(classes_dir) if f.endswith('.csv')]
    if not class_csv_files:
        print_admin_warning(f"Warnung: Keine Klassen-CSV-Dateien im Ordner '{classes_dir}' für die Erstellung von Admin-Warnungen gefunden.")
        return admin_warnings_cache  # Gibt leere Warnungsliste zurück oder handle dies mit Dummy-Daten

    # Neueste Klassen-CSV-Datei bestimmen
    newest_class_file = max(class_csv_files, key=lambda f: os.path.getctime(os.path.join(classes_dir, f)))

    # Warnungen für fehlende Klassenlehrkräfte in der Klassen-Datei (Spalten 8 und 9)
    with open(os.path.join(classes_dir, newest_class_file), 'r', newline='', encoding='utf-8-sig') as class_file:
        class_reader = csv.reader(class_file, delimiter=';')
        header = next(class_reader)
        for row in class_reader:
            for idx in [7, 8]:  # Indizes für Klassenlehrkräfte
                teacher_name = row[idx].strip()
                if teacher_name and teacher_name not in teacher_names:
                    admin_warnings_cache.append({
                        'Typ': 'Fehlender Klassenlehrer in Lehrerdatei',
                        'Details': f"Der Klassenlehrer '{teacher_name}' aus der Klassen-Datei existiert nicht in der Lehrkräfte-Datei."
                    })

    # Admin-Warnungen in der Konsole ausgeben
    if admin_warnings_cache:
        print_admin_warning("Admin-Warnungen:")
        for warning in admin_warnings_cache:
            print_admin_warning(f"Typ: {warning['Typ']}, Details: {warning['Details']}")
        print_warningtext("Für die korrekte Funktion Warn-Funktionalität sollten Sie Ihre Klassen- und Lehrerdaten aktualisieren.")
    else:
        print_success("Keine Admin-Warnungen gefunden.")

    # E-Mail senden, falls das Kommandozeilenargument verwendet wurde
    if send_email_flag and admin_warnings_cache:
        config.read('email_settings.ini', encoding='utf-8-sig')
        admin_email = config.get('Email', 'admin_email', fallback=None)
        if not admin_email:
            print_error("Admin-E-Mail-Adresse fehlt in email_settings.ini.")
            return admin_warnings_cache

        subject = "Admin-Warnungen von WebUntis"
        body = "<p>Folgende Admin-Warnungen wurden generiert:</p><ul>"
        for warning in admin_warnings_cache:
            body += f"<li><strong>{warning['Typ']}:</strong> {warning['Details']}</li>"
        body += "</ul><p>Mit freundlichen Grüßen,<br>Ihr System</p>"

        try:
            send_email(subject, body, [admin_email])
            print_success(f"Admin-Warnungen wurden erfolgreich an {admin_email} gesendet.")
        except Exception as e:
            print_error(f"Fehler beim Senden der Admin-Warnungen: {e}")

    return admin_warnings_cache


#(Get) Route und Funktion zum Öffnen der Webseite, Reinladen der Daten, Initialisieren von Caches für Warnungen, Fehlermeldungen etc., 
#(Post) sowie auch Ausführung der def_run aus der main.py bei Klick auf den Verarbeiten-Button
@app.route('/test')
def test_route(): return 'TEST OK', 200

@app.route('/test_api', methods=['POST'])
def test_api():
    from webuntis_api import WebUntisClient
    data = request.json
    client = WebUntisClient(
        server_url=data.get('server_url'),
        school=data.get('school'),
        user=data.get('user'),
        password=data.get('password'),
        client_name=data.get('client_name', 'Schild-WebUntis-Tool')
    )
    success, message = client.test_connection()
    return jsonify({'success': success, 'message': message})

@app.route('/', methods=['GET', 'POST'])
def index():
    global warnings_cache  # Globaler Cache für die Zwischenspeicherung von Warnungen
    warnings = warnings_cache or []
    confirmation = None  # Variable für Bestätigungsnachricht
    errors = []  # Liste für Fehlermeldungen
    warnings_messages = []  # Liste für nicht-blockierende Warnungen

    # Zugriff auf die globalen CLI-Argumente
    global cli_args
    no_log = cli_args.get("no_log", False)
    no_xlsx = cli_args.get("no_xlsx", False)

    # Werte aus der settings.ini laden
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    use_abschlussdatum = config.getboolean('ProcessingOptions', 'use_abschlussdatum', fallback=False)
    create_second_file = config.getboolean('ProcessingOptions', 'create_second_file', fallback=False)
    warn_entlassdatum = config.getboolean('ProcessingOptions', 'warn_entlassdatum', fallback=True)
    warn_aufnahmedatum = config.getboolean('ProcessingOptions', 'warn_aufnahmedatum', fallback=True)
    warn_klassenwechsel = config.getboolean('ProcessingOptions', 'warn_klassenwechsel', fallback=True)
    class_change_recipients = config.get('ProcessingOptions', 'class_change_recipients', fallback='old')
    warn_new_students = config.getboolean('ProcessingOptions', 'warn_new_students', fallback=True)
    warn_karteileichen = config.getboolean('ProcessingOptions', 'warn_karteileichen', fallback=False)
    create_class_size_file = config.getboolean('ProcessingOptions', 'create_class_size_file', fallback=True) 
    enable_attestpflicht_column = config.getboolean('ProcessingOptions', 'enable_attestpflicht_column', fallback=False) 
    enable_nachteilsausgleich_column= config.getboolean('ProcessingOptions', 'enable_nachteilsausgleich_column', fallback=False)
    disable_import_file_creation = config.getboolean('ProcessingOptions', 'disable_import_file_creation', fallback=False)
    disable_import_file_if_admin_warning = config.getboolean('ProcessingOptions', 'disable_import_file_if_admin_warning', fallback=False)

    # Werte aus der email_settings.ini laden
    config = configparser.ConfigParser()
    config.read("email_settings.ini", encoding='utf-8-sig')

    # Vorlagenwerte laden
    subject_entlassdatum = config.get("Templates", "subject_entlassdatum", fallback="")
    body_entlassdatum = config.get("Templates", "body_entlassdatum", fallback="")
    subject_aufnahmedatum = config.get("Templates", "subject_aufnahmedatum", fallback="")
    body_aufnahmedatum = config.get("Templates", "body_aufnahmedatum", fallback="")
    subject_klassenwechsel = config.get("Templates", "subject_klassenwechsel", fallback="")
    body_klassenwechsel = config.get("Templates", "body_klassenwechsel", fallback="")
    subject_karteileiche = config.get("Templates", "subject_karteileiche", fallback="")
    body_karteileiche = config.get("Templates", "body_karteileiche", fallback="")

    # Einstellungen laden
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    classes_dir = config.get("Directories", "classes_directory", fallback="./Klassendaten")
    teachers_dir = config.get("Directories", "teachers_directory", fallback="./Lehrerdaten")

    # Überprüfen, ob die Haupt-CSV-Datei vorhanden ist
    schildexport_dir = get_directory('schildexport_directory', default='.')
    if schildexport_dir in ('.', '', None):
        schildexport_dir = os.getcwd()
    main_csv_exists = any(
        f.endswith('.csv') for f in os.listdir(schildexport_dir) if not os.path.isdir(os.path.join(schildexport_dir, f))
    )

    if not main_csv_exists:
        errors.append("Die Haupt-CSV-Datei fehlt im Hauptverzeichnis und wird für die Verarbeitung benötigt.")
        print_error("Fehler: Haupt-CSV-Datei fehlt im Hauptverzeichnis und wird für die Verarbeitung benötigt.")

    # Überprüfen, ob die Klassendaten verfügbar sind
    if not os.path.exists(classes_dir) or not any(f.endswith('.csv') for f in os.listdir(classes_dir)):
        warnings_messages.append("Die Klassendaten fehlen oder es sind keine CSV-Dateien im konfigurierten Ordner vorhanden.")
        print_warning(f"Warnung: Keine Klassendaten im Ordner '{classes_dir}' zur Vorbereitung der Warnungen gefunden.")

    # Überprüfen, ob die Lehrerdaten verfügbar sind
    if not os.path.exists(teachers_dir) or not any(f.endswith('.csv') for f in os.listdir(teachers_dir)):
        warnings_messages.append("Die Lehrerdaten fehlen oder es sind keine CSV-Dateien im konfigurierten Ordner vorhanden.")
        print_warning(f"Warnung: Keine Lehrerdaten im Ordner '{teachers_dir}' zur Vorbereitung der Warnungen gefunden.")

    if request.method == 'POST' and not errors:
        # Aktuelle Werte aus dem Formular auf im WebEnd lesen und die Auswahl des Benutzers speichern (Standardwerte werden für den Prozess überschrieben)
        use_abschlussdatum = request.form.get('use_abschlussdatum') == 'on'
        create_second_file = request.form.get('create_second_file') == 'on'
        warn_entlassdatum = request.form.get('warn_entlassdatum') == 'on'
        warn_aufnahmedatum = request.form.get('warn_aufnahmedatum') == 'on'
        warn_klassenwechsel = request.form.get('warn_klassenwechsel') == 'on'
        class_change_recipients = request.form.get('class_change_recipients', class_change_recipients)
        warn_new_students = request.form.get('warn_new_students') == 'on'
        warn_karteileichen = request.form.get('warn_karteileichen') == 'on'
        create_class_size_file = request.form.get('create_class_size_file') == 'on'
        enable_attestpflicht_column = request.form.get('enable_attestpflicht_column') == 'on'
        enable_nachteilsausgleich_column = request.form.get('enable_nachteilsausgleich_column') == 'on'
        disable_import_file_creation = request.form.get('disable_import_file_creation') == 'on'
        disable_import_file_if_admin_warning = request.form.get('disable_import_file_if_admin_warning') == 'on'

        # Übergebe die Auswahl an die Run-Funktion
        try:
            # Datenverarbeitung basierend auf den Benutzereinstellungen
            print_info("🌐 Dashboard-Aktion: Manueller Daten-Import durch Benutzer gestartet.")
            print("-" * 40)
            print_info("Starte Verarbeitung über die Weboberfläche...")
            warnings = run(
                use_abschlussdatum=use_abschlussdatum,
                create_second_file=create_second_file,
                warn_entlassdatum=warn_entlassdatum,
                warn_aufnahmedatum=warn_aufnahmedatum,
                warn_klassenwechsel=warn_klassenwechsel,
                warn_new_students=warn_new_students,
                warn_karteileichen=warn_karteileichen,
                class_change_recipients=class_change_recipients,
                no_log=no_log,
                no_xlsx=no_xlsx,
                create_class_size_file=create_class_size_file, 
                enable_attestpflicht_column=enable_attestpflicht_column,
                enable_nachteilsausgleich_column=enable_nachteilsausgleich_column,
                disable_import_file_creation=disable_import_file_creation,
                disable_import_file_if_admin_warning=disable_import_file_if_admin_warning,
                admin_warnings_cache=admin_warnings_cache
            )
            for w in warnings:
                w['status'] = 'offen'
            warnings_cache = warnings  # Speichert Warnungen zur späteren Nutzung

            # Setzt eine Bestätigungsnachricht nach erfolgreicher Ausführung
            confirmation = "Verarbeitung erfolgreich abgeschlossen."
            print_success("Verarbeitung über die Weboberfläche erfolgreich abgeschlossen.")
        except Exception as e:
            errors.append(f"Fehler während der Verarbeitung: {str(e)}")
            print_error(f"Fehler während der Verarbeitung über die Weboberfläche: {str(e)}")
        print("-" * 40)

    # Seite mit aktuellen Checkbox-Zuständen, Warnungen, Fehlern und Bestätigung rendern
    return render_template(
        'index.html',
        warnings=warnings,
        admin_warnings=admin_warnings_cache,
        confirmation=confirmation,
        errors=errors,
        warnings_messages=warnings_messages,
        use_abschlussdatum=use_abschlussdatum,
        create_second_file=create_second_file,
        warn_entlassdatum=warn_entlassdatum,
        warn_aufnahmedatum=warn_aufnahmedatum,
        warn_klassenwechsel=warn_klassenwechsel,
        class_change_recipients=class_change_recipients,
        warn_new_students=warn_new_students,
        warn_karteileichen=warn_karteileichen,
        create_class_size_file=create_class_size_file,
        enable_attestpflicht_column = enable_attestpflicht_column,
        enable_nachteilsausgleich_column = enable_nachteilsausgleich_column,
        disable_import_file_creation=disable_import_file_creation,
        disable_import_file_if_admin_warning=disable_import_file_if_admin_warning,
        subject_entlassdatum=subject_entlassdatum,
        body_entlassdatum=body_entlassdatum,
        subject_aufnahmedatum=subject_aufnahmedatum,
        body_aufnahmedatum=body_aufnahmedatum,
        subject_klassenwechsel=subject_klassenwechsel,
        body_klassenwechsel=body_klassenwechsel,
        subject_karteileiche=subject_karteileiche,
        body_karteileiche=body_karteileiche,
        no_directory_change=cli_args.get("no_directory_change", False),
        enable_upload=cli_args.get("enable_upload", False),
        initial_validation=validate_imports()
    )


from string import Template

# Route und Funktion hinter dem Button "E-Mails Generieren". Sie generiert die E-Mails zur Vorschau im WebEnd (view_emails.html).
@app.route('/generate_emails', methods=['POST'])
def generate_emails():
    global warnings_cache, generated_emails_cache
    generated_emails_cache = []  # Globaler Cache für generierte E-Mails

    if warnings_cache:
        print_info("🌐 Dashboard-Aktion: E-Mail-Generierung angefordert.")
        print_info("Generiere E-Mails basierend auf den vorhandenen Warnungen...")
        # E-Mail-Einstellungen laden
        config = configparser.ConfigParser()
        safe_read_config(config, 'email_settings.ini')

        for i, warning in enumerate(warnings_cache):
            # Bestimme den Typ der Warnung
            if 'neues_entlassdatum' in warning:
                warning_type = "entlassdatum"
            elif 'neues_aufnahmedatum' in warning:
                warning_type = "aufnahmedatum"
            elif 'neue_klasse' in warning:
                warning_type = "klassenwechsel"
            elif 'new_student' in warning and warning['new_student']:
                warning_type = "new_student"
            elif 'karteileiche' in warning and warning['karteileiche']:
                warning_type = "karteileiche"
            else:
                continue  # Unbekannter Warnungstyp wird übersprungen

            # Dynamische Einbindung des Zeitraums (falls vorhanden)
            zeitraum_text = (
                f"<p><strong>Zeitraum nicht dokumentiert:</strong> {warning.get('Zeitraum_nicht_dokumentiert', 'N/A')}</p>"
                if 'Zeitraum_nicht_dokumentiert' in warning
                else ""
            )

            # Prüfen, ob "Volljaehrig" verwendet wird, und den Wert dynamisch bereitstellen
            if 'Volljährig' not in warning:
                warning['Volljaehrig'] = warning.get('Volljaehrig', 'Unbekannt')

            # Lade Vorlagen aus der .ini-Datei
            try:
                subject_template = config.get("Templates", f"subject_{warning_type}")
                body_template = config.get("Templates", f"body_{warning_type}")
            except configparser.NoOptionError:
                error_message = f"Vorlage für {warning_type} fehlt in der Konfigurationsdatei."
                print_error(error_message)
                return jsonify({"message": f"⚠️ Vorlage für {warning_type} fehlt in der Konfigurationsdatei."}), 400

            # Verwende Template-System zur Verarbeitung der Vorlagen
            try:
                subject = Template(subject_template).substitute(**warning)
                body = Template(body_template).substitute(
                    **warning,
                    zeitraum_text=zeitraum_text  # Zusatzwert für Zeitraum
                )
            except KeyError as e:
                error_message = f"Fehlender Platzhalter: {e} in der Vorlage für {warning_type}"
                print_error(error_message)
                return jsonify({"message": f"⚠️ Fehlender Platzhalter: {e} in der Vorlage für {warning_type}"}), 400

            # E-Mail zur Liste hinzufügen
            recipients = warning.get('recipients_list')
            if not recipients:
                recipients = [warning.get('Klassenlehrkraft_1_Email', 'N/A'), warning.get('Klassenlehrkraft_2_Email', 'N/A')]

            generated_emails_cache.append({
                'subject': subject,
                'body': body,
                'to': recipients,
                'warning_index': i
            })
        print_success("E-Mails wurden erfolgreich generiert.")
        return jsonify({"message": "✅ Die E-Mails wurden erfolgreich generiert.", "emails": generated_emails_cache})
    else:
        print_info("Keine Warnungen vorhanden, um E-Mails zu generieren.")
        return jsonify({"message": "ℹ️ Keine Warnungen verfügbar, um E-Mails zu generieren."})

# Abruf der generierten E-Mails im, WebEnd mittels der view_emails.html
@app.route('/view_generated_emails', methods=['GET'])
def view_generated_emails():
    global generated_emails_cache
    print_info("Anzeige der generierten E-Mails im Webinterface...")
    # Rendert die Seite zum Anzeigen der generierten E-Mails
    return render_template('view_emails.html', emails=generated_emails_cache)

# Route zum Abrufen der Dateihistorie (Logs & Excels)
@app.route('/api/history', methods=['GET'])
def get_history():
    print_info("Lade Historie-Daten...")
    config = configparser.ConfigParser()
    config.read("settings.ini", encoding='utf-8-sig')
    log_dir = config.get("Directories", "log_directory", fallback="Logs")
    xlsx_dir = config.get("Directories", "xlsx_directory", fallback="ExcelExports")

    history_dict = {}
    
    import re
    date_pattern = re.compile(r'_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.')

    def scan_dir(directory, extension, key_name):
        if not os.path.exists(directory): return
        for f in os.listdir(directory):
            if f.endswith(extension):
                match = date_pattern.search(f)
                if match:
                    timestamp = match.group(1)
                    if timestamp not in history_dict:
                        # Format timestamp for display nicely
                        display_time = timestamp.replace('_', ' ').replace('-', ':', 2)
                        # Quick fix: original format is YYYY-MM-DD_HH-MM-SS
                        # We want YYYY-MM-DD HH:MM:SS
                        parts = timestamp.split('_')
                        if len(parts) == 2:
                            display_time = f"{parts[0]} {parts[1][:2]}:{parts[1][3:5]}:{parts[1][6:]}"

                        history_dict[timestamp] = {
                            "timestamp": timestamp,
                            "display_time": display_time,
                            "log_file": None,
                            "xlsx_file": None
                        }
                    history_dict[timestamp][key_name] = f

    scan_dir(log_dir, '.log', 'log_file')
    scan_dir(xlsx_dir, '.xlsx', 'xlsx_file')

    # Convert to list and sort descending by timestamp
    history_list = list(history_dict.values())
    history_list.sort(key=lambda x: x["timestamp"], reverse=True)

    # Apply limit
    limit = request.args.get('limit', default=30, type=int)
    if limit > 0:
        history_list = history_list[:limit]

    return jsonify(history_list)


# Route zum Anzeigen des Datei-Inhalts einer .log-Datei im WeUI Modal
@app.route('/api/log_content/<path:filename>', methods=['GET'])
def get_log_content(filename):
    config = configparser.ConfigParser()
    config.read("settings.ini", encoding='utf-8-sig')
    log_dir = config.get("Directories", "log_directory", fallback="Logs")
    
    # Path-Traversal Protection
    safe_filename = os.path.basename(filename)
    file_path = os.path.join(log_dir, safe_filename)

    if not os.path.exists(file_path):
        return jsonify({"error": "Datei nicht gefunden"}), 404
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except Exception as e:
        # Fallback falls es anderes Encoding hat
        try:
            with open(file_path, 'r', encoding='utf-16') as f:
                content = f.read()
                return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        except Exception as e2:
            return jsonify({"error": f"Dateiformat kann nicht gelesen werden: {str(e2)}"}), 500

# Route zum Anzeigen des Inhalts einer Excel-Log-Datei im WeUI Modal (Konvertierung in HTML-Tabelle)
@app.route('/api/xlsx_view/<path:filename>', methods=['GET'])
def view_xlsx(filename):
    from openpyxl import load_workbook
    config = configparser.ConfigParser()
    config.read("settings.ini", encoding='utf-8-sig')
    xlsx_dir = config.get("Directories", "xlsx_directory", fallback="ExcelExports")

    # Pfad-Sicherheitsprüfung
    safe_filename = os.path.basename(filename)

    # Pfad absolut machen
    if not os.path.isabs(xlsx_dir):
        xlsx_dir = os.path.abspath(xlsx_dir)
    file_path = os.path.join(xlsx_dir, safe_filename)

    if not os.path.exists(file_path):
        return jsonify({"error": "Excel-Datei nicht gefunden"}), 404

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # HTML-Tabelle generieren
        html = '<div class="table-responsive"><table class="table table-sm table-bordered table-striped">'
        for row in ws.iter_rows(values_only=True):
            html += '<tr>'
            for cell in row:
                # Zellinhalt sicher als String behandeln
                cell_val = str(cell) if cell is not None else ""
                # Stil wie im Excel (rote Markierung bei "->")
                style = ' style="background-color: #FFCCCC;"' if "->" in cell_val else ""
                html += f'<td{style}>{cell_val}</td>'
            html += '</tr>'
        html += '</table></div>'

        return html, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        return jsonify({"error": f"Excel konnte nicht gelesen werden: {str(e)}"}), 500


@app.route('/api/refresh_admin_warnings', methods=['POST'])
def refresh_admin_warnings():
    print_info("🌐 Dashboard-Aktion: Admin-Warnungen (System-Check) manuell neu gestartet.")
    global admin_warnings_cache
    admin_warnings_cache.clear()
    admin_warnings()
    return jsonify({"success": True, "count": len(admin_warnings_cache), "warnings": admin_warnings_cache})

# Route zum Herunterladen einer Excel-Log-Datei aus dem ExcelExports-Verzeichnis
@app.route('/api/xlsx_download/<path:filename>', methods=['GET'])
def download_xlsx(filename):
    config = configparser.ConfigParser()
    config.read("settings.ini", encoding='utf-8-sig')
    xlsx_dir = config.get("Directories", "xlsx_directory", fallback="ExcelExports")
    
    # Pfad-Sicherheitsprüfung
    safe_filename = os.path.basename(filename)
    # Sicherstellen, dass das Verzeichnis absolut ist, falls nötig
    if not os.path.isabs(xlsx_dir):
        xlsx_dir = os.path.abspath(xlsx_dir)
        
    if not os.path.exists(os.path.join(xlsx_dir, safe_filename)):
        return jsonify({"error": "Excel-Datei nicht gefunden"}), 404
        
    return send_from_directory(xlsx_dir, safe_filename, as_attachment=True)
# Route und Funktion zum Abruf der E-Mail Inhalte des Vorlagen-Email-Editors im WebEnd. 
@app.route('/get_templates', methods=['GET'])
def get_templates():
    print_info("Lade bzw. Aktualisiere die Inhalte der Weboberfläche:")
    # Lädt die E-Mail-Vorlagen aus der Konfigurationsdatei
    print_info("Lade E-Mail-Vorlagen aus 'email_settings.ini' für den E-Mail-Vorlagen Editor...")
    config = configparser.ConfigParser()
    try:
        # E-Mail-Vorlagen laden
        config.read('email_settings.ini', encoding='utf-8-sig')

        # Rückgabeobjekt vorbereiten
        templates = {
            "subject_entlassdatum": config.get("Templates", "subject_entlassdatum", fallback=""),
            "body_entlassdatum": config.get("Templates", "body_entlassdatum", fallback=""),
            "subject_aufnahmedatum": config.get("Templates", "subject_aufnahmedatum", fallback=""),
            "body_aufnahmedatum": config.get("Templates", "body_aufnahmedatum", fallback=""),
            "subject_klassenwechsel": config.get("Templates", "subject_klassenwechsel", fallback=""),
            "body_klassenwechsel": config.get("Templates", "body_klassenwechsel", fallback=""),
            "subject_new_student": config.get("Templates", "subject_new_student", fallback=""),
            "body_new_student": config.get("Templates", "body_new_student", fallback=""),
            "subject_karteileiche": config.get("Templates", "subject_karteileiche", fallback=""),
            "body_karteileiche": config.get("Templates", "body_karteileiche", fallback="")
        }

        # Erfolg erst melden, wenn sicher ist, dass alles funktioniert hat
        print_success("E-Mail-Vorlagen erfolgreich in den E-Mail-Vorlagen Editor geladen.")
        return jsonify(templates)

    except Exception as e:
        # Fehlerfall behandeln und Fehler ausgeben
        error_message = f"Fehler beim Laden der E-Mail-Vorlagen: {str(e)}"
        print_error(error_message)
        return jsonify({"error": str(e)}), 500

# Route und Funktion zum Speichern der Inhalte des Vorlagen-Email-Editors im WebEnd. 
@app.route('/update_templates', methods=['POST'])
def update_templates():
    # Aktualisiert die E-Mail-Vorlagen basierend auf den vom Benutzer gesendeten Daten
    print_info("Aktualisiere E-Mail-Vorlagen in 'email_settings.ini'...")
    try:
        # E-Mail-Einstellungen laden
        email_config = configparser.ConfigParser()
        safe_read_config(email_config, 'email_settings.ini')

        # Vorlagen mit den bereitgestellten Daten aktualisieren
        email_config['Templates']['subject_entlassdatum'] = request.form.get('subject_entlassdatum', '')
        email_config['Templates']['body_entlassdatum'] = request.form.get('body_entlassdatum', '')
        email_config['Templates']['subject_aufnahmedatum'] = request.form.get('subject_aufnahmedatum', '')
        email_config['Templates']['body_aufnahmedatum'] = request.form.get('body_aufnahmedatum', '')
        email_config['Templates']['subject_klassenwechsel'] = request.form.get('subject_klassenwechsel', '')
        email_config['Templates']['body_klassenwechsel'] = request.form.get('body_klassenwechsel', '')
        email_config['Templates']['subject_new_student'] = request.form.get('subject_new_student', '')
        email_config['Templates']['body_new_student'] = request.form.get('body_new_student', '')
        email_config['Templates']['subject_karteileiche'] = request.form.get('subject_karteileiche', '')
        email_config['Templates']['body_karteileiche'] = request.form.get('body_karteileiche', '')

        # Änderungen in die Datei schreiben
        with open('email_settings.ini', 'w', encoding='utf-8-sig') as configfile:
            email_config.write(configfile)
        print_success("E-Mail-Vorlagen wurden erfolgreich aktualisiert.")
        return jsonify({'message': '✅ E-Mail-Vorlagen erfolgreich gespeichert!'})
    except Exception as e:
        error_message = f'Fehler beim Speichern der E-Mail-Vorlagen: {str(e)}'
        print_error(error_message)
        return jsonify({'message': f'❌ Fehler beim Speichern der E-Mail-Vorlagen: {str(e)}'}), 500

@app.route('/api/templates/default/<template_type>', methods=['GET'])
def get_default_template(template_type):
    if template_type in DEFAULT_TEMPLATES:
        return jsonify(DEFAULT_TEMPLATES[template_type])
    return jsonify({"error": "Template type not found"}), 404


# Route und Funktion hinter dem "E-Mails Senden" Button im WebEnd zum Senden der E-Mails auf Grundlage der generierten Warnugen und gespeicherten Einstellungen 
@app.route('/send_emails', methods=['POST'])
def send_emails():
    global generated_emails_cache, warnings_cache
    if generated_emails_cache:
        print_info("Beginne mit dem Senden der generierten E-Mails...")
        # Verwende den Cache zum Senden der E-Mails
        for email in generated_emails_cache:
            # Filtere N/A Adressen vor dem Senden
            actual_recipients = [r for r in email['to'] if r and r.lower() != 'n/a']
            
            if not actual_recipients:
                print_warning(f"Überspringe E-Mail für '{email['subject']}', da keine gültigen Empfänger vorhanden sind.")
                continue

            try:
                send_email(
                    subject=email['subject'],
                    body=email['body'],
                    to_addresses=actual_recipients
                )
                print_success(f"✅ E-Mail an {actual_recipients} erfolgreich gesendet.")
                
                # Status der Warnung aktualisieren
                idx = email.get('warning_index')
                if idx is not None and idx < len(warnings_cache):
                    warnings_cache[idx]['status'] = 'versendet'
                
                # Kurze Pause um Rate-Limiting zu vermeiden (z.B. bei Strato)
                import time
                time.sleep(2)
            except Exception as e:
                print_error(f"❌ ❌ ❌ Fehler beim Senden an {actual_recipients}: {str(e)}")
                
        print_success("--- Alle E-Mails wurden verarbeitet ---")
        return jsonify({"message": "📧 Die Verarbeitung der E-Mails ist abgeschlossen."})
    else:
        print_warning("Keine generierten E-Mails zum Senden vorhanden.")
        return jsonify({"message": " ⚠️Keine generierten E-Mails verfügbar, um sie zu versenden."})

# API-Route zum Abrufen der aktuellen Warnungen mit ihrem Status
@app.route('/api/get_warnings', methods=['GET'])
def get_warnings():
    global warnings_cache
    return jsonify(warnings_cache)

# API-Route zum Durchführen eines Vorab-Checks der Import-Dateien
@app.route('/api/validate_imports', methods=['GET'])
def api_validate_imports():
    report = validate_imports()
    return jsonify(report)

# API-Route zum Prüfen, ob bereits E-Mails generiert wurden
@app.route('/api/check_emails_status', methods=['GET'])
def check_emails_status():
    global generated_emails_cache
    return jsonify({"has_emails": len(generated_emails_cache) > 0})

# --- Dashboard & Historie API ---

@app.route('/api/history/stats', methods=['GET'])
def get_history_stats():
    field_filter = request.args.get('field')
    stats = history_manager.get_dashboard_stats(field_filter=field_filter)
    return jsonify(stats)

@app.route('/api/history/classes', methods=['GET'])
def get_history_classes():
    classes = history_manager.get_all_classes()
    return jsonify(classes)

@app.route('/api/history/class_stats/<class_name>', methods=['GET'])
def get_class_stats(class_name):
    stats = history_manager.get_class_analytics(class_name)
    return jsonify(stats)

@app.route('/api/history/export/excel', methods=['GET'])
def export_history_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    
    export_type = request.args.get('type') # 'student' or 'class'
    identifier = request.args.get('id')
    
    if not export_type or not identifier:
        return jsonify({"error": "Missing parameters"}), 400
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Historie"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if export_type == 'student':
        data = history_manager.search_student_history(identifier)
        if not data: return "No data", 404
        student = data[0]['student']
        timeline = data[0]['timeline']
        ws.append(["Datum", "Feld", "Alter Wert", "Neuer Wert", "Quelle"])
        for t in timeline:
            ws.append([t['timestamp'], t['field'], t['old_value'], t['new_value'], t['file_b']])
        filename = f"Historie_{student['name']}_{identifier}.xlsx"
    else:
        stats = history_manager.get_class_analytics(identifier)
        timeline = stats['timeline']
        ws.append(["Datum", "Schüler Name", "Schüler ID", "Feld", "Alter Wert", "Neuer Wert"])
        for t in timeline:
            ws.append([t['timestamp'], t['name'], t['student_id'], t['field'], t['old_value'], t['new_value']])
        filename = f"Klassen_Historie_{identifier}.xlsx"

    # Spaltenbreite und Styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/history/sync', methods=['POST'])
def sync_history_classes():
    import main
    success = main.sync_student_classes_from_latest()
    if success:
        return jsonify({"success": True, "message": "Synchronisierung erfolgreich abgeschlossen."})
    else:
        return jsonify({"success": False, "message": "Synchronisierung fehlgeschlagen oder keine Quelldaten gefunden."}), 500

@app.route('/api/history/search', methods=['GET'])
def search_history():
    query = request.args.get('q', '')
    if len(query) < 2:
        return jsonify([])
    results = history_manager.search_student_history(query)
    return jsonify(results)

@app.route('/api/history/reset', methods=['POST'])
def reset_history():
    history_manager.clear_history()
    return jsonify({"message": "✅ Historie erfolgreich zurückgesetzt."})

@app.route('/api/history/reindex', methods=['POST'])
def reindex_history():
    config = configparser.ConfigParser()
    from utils import safe_read_config
    safe_read_config(config, 'settings.ini')
    log_dir = config.get("Directories", "log_directory", fallback="./Logs")
    count = history_manager.reindex_logs(log_dir)
    if count == 0:
        msg = "ℹ️ Keine neuen Log-Dateien gefunden (alle bereits importiert oder Verzeichnis leer)."
    else:
        msg = f"✅ {count} neue Log-Datei(en) erfolgreich in die Datenbank importiert."
    return jsonify({"message": msg})

# Route und Funktion zum Abrufen und Reinladen der Einstellungen des Einstellungs-Panels im WebEnd aus den verschiedenen .ini Dateien
@app.route('/load-settings', methods=['GET'])
def load_settings():
    print_info("Lade Einstellungen aus 'settings.ini' und 'email_settings.ini' für das Einstellungen-Panel...")
    settings = {}
    # settings.ini laden
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    for section in config.sections():
        if section not in settings:
            settings[section] = {}
        settings[section].update({key: config.get(section, key, fallback="") for key in config[section]})

    # email_settings.ini laden
    email_config = configparser.ConfigParser()
    safe_read_config(email_config, "email_settings.ini")
    for section in email_config.sections():
        if section not in settings:
            settings[section] = {}
        settings[section].update({key: email_config.get(section, key, fallback="") for key in email_config[section]})
    print_success("Einstellungen für das Einstellungs-Panel erfolgreich geladen bzw. aktualisiert.")
    return jsonify(settings)

# Route und Funktion zum Speichern der geänderten Einstellungen aus dem Einstellungs-Panel im WebEnd in die verschiedenen .ini Dateien. 
# Hier wird zunächst sortiert, welche Daten in welche Datei gehören.
@app.route('/save-settings', methods=['POST'])
def save_settings():
    print_info("Speichere Einstellungen in 'settings.ini' und 'email_settings.ini'...")
    settings = request.json.get('settings', {})
    settings_ini_data = {}
    email_settings_ini_data = {}

    # Definiere, welche Abschnitte zu email_settings.ini gehören
    email_sections = ['Email', 'OAuth', 'Templates']

    # Aufteilen der Einstellungen in die entsprechenden Dateien
    for section, values in settings.items():
        if section in email_sections:
            if section not in email_settings_ini_data:
                email_settings_ini_data[section] = {}
            email_settings_ini_data[section].update(values)
        else:
            if section not in settings_ini_data:
                settings_ini_data[section] = {}
            settings_ini_data[section].update(values)

    # If directory change is disabled, remove the Directories section or ignore changes to directories
    if cli_args.get("no_directory_change", False):
        if 'Directories' in settings_ini_data:
            print_info("Directory changing is disabled via command-line argument. Ignoring changes to directories.")
            del settings_ini_data['Directories']

    # Einstellungen speichern
    if settings_ini_data:
        save_to_settings_ini(settings_ini_data) #Speichern der für in die settings.ini gehörenden Daten
    if email_settings_ini_data:
        save_to_email_settings_ini(email_settings_ini_data) #Speichern der für in die email_settings.ini gehörenden Daten

    print_success("Einstellungen wurden erfolgreich gespeichert.")
    return jsonify({"status": "success"})


# Funktion zum Speichern der Einstellungen in die Datei 'settings.ini'
def save_to_settings_ini(settings):
    print_info("Speichere Einstellungen in 'settings.ini'...")
    config = configparser.ConfigParser()
    safe_read_config(config, "settings.ini")
    for section, values in settings.items():
        if not config.has_section(section):
            config.add_section(section)
            print_info(f"Abschnitt '{section}' hinzugefügt.")
        for key, value in values.items():
            # Ersetze Backslashes durch Forward Slashes für Pfade
            if "directory" in key:
                value = os.path.normpath(value)
                print_info(f"Pfad normalisiert für '{key}': {value}")
            config.set(section, key, str(value))
            print_info(f"Einstellung gesetzt: [{section}] {key} = {value}")
    with open("settings.ini", "w", encoding="utf-8-sig") as configfile:
        config.write(configfile)
    print_success("Einstellungen wurden erfolgreich in 'settings.ini' gespeichert.")

# Funktion zum Speichern der E-Mail-Einstellungen in 'email_settings.ini'
def save_to_email_settings_ini(settings):
    print_info("Speichere E-Mail-Einstellungen in 'email_settings.ini'...")
    config = configparser.ConfigParser()
    safe_read_config(config, "email_settings.ini")
    # Bestehende Abschnitte beibehalten
    existing_sections = config.sections()
    # Aktualisiere nur die bereitgestellten Abschnitte
    for section, values in settings.items():
        print_info(f"Aktualisiere Abschnitt '{section}'...")
        if section in existing_sections:
            for key, value in values.items():
                config.set(section, key, str(value))
                print_info(f"Einstellung aktualisiert: [{section}] {key} = {value}")
        else:
            print_info(f"Füge neuen Abschnitt '{section}' hinzu...")
            config[section] = values
            for key, value in values.items():
                print_info(f"Einstellung hinzugefügt: [{section}] {key} = {value}")
    # Einstellungen in die Datei schreiben
    with open("email_settings.ini", "w", encoding="utf-8-sig") as configfile:
        config.write(configfile)
    print_success("E-Mail-Einstellungen wurden erfolgreich in 'email_settings.ini' gespeichert.")


# Route und Funktion für den Durchsuchen-Button zur Angabe und Speicherung vollständiger absoluter Verzeichnispfade
@app.route('/process-directory', methods=['POST'])
def process_directory():
    print_info("Verarbeite ausgewähltes Verzeichnis...")
    directory_name = request.form.get("directoryName")
    if not directory_name:
        error_message = "Verzeichnisname ist erforderlich."
        print_error(f"Fehler: {error_message}")
        return jsonify({"error": error_message}), 400

    # Beispiel: Verarbeite das Verzeichnis (Anpassung an Ihr System erforderlich)
    full_path = os.path.abspath(directory_name)

    # Konvertiere Backslashes zu Forward Slashes für INI-Kompatibilität
    formatted_path = full_path.replace("\\", "/")

    return jsonify({"fullPath": formatted_path})

# Route und Funktion für den Durchsuchen-Button zur Angabe und Speicherung vollständiger absoluter Verzeichnispfade
@app.route('/select-directory', methods=['POST'])
def select_directory():
    print_info("Öffne Dateiauswahl zur Auswahl eines Verzeichnisses...")

    # Variable zum Speichern des Ergebnisses
    result = {'directory': None}
    
    def open_dialog():
        try:
            root = tk.Tk()
            root.withdraw()  # Hauptfenster ausblenden
            root.wm_attributes('-topmost', 1)  # Setzt das Fenster in den Vordergrund
            directory = filedialog.askdirectory()
            root.destroy()
            if directory:
                print_info(f"Verzeichnis ausgewählt: {directory}")
                result['directory'] = directory
            else:
                print_warning("Keine Auswahl getroffen.")
        except Exception as e:
            error_message = f"Fehler beim Öffnen der Dateiauswahl: {str(e)}"
            print_error(error_message)

    thread = threading.Thread(target=open_dialog)
    thread.start()
    thread.join()

    if result['directory']:
        return jsonify({"selected_directory": result['directory']})
    else:
        return jsonify({"selected_directory": None})


# Route und Funktion zum Abrufen der verfügbaren Kommandozeilenargumente mit ihren Beschreibungen, um die Argumente für den Befehl- und Verknüpfungsersteller bereitzustellen
@app.route('/get-arguments', methods=['GET'])
def get_arguments():

    print_info(f"Bereite Liste der verfügbaren Kommandozeilenargumente für das Befehl- und Verknüpfungs-Erstelltool vor...")
    # Liste der Argumente mit Beschreibungen
    arguments = [
        {"name": "--process", "description": "Führt den Hauptprozess aus (Verarbeitung des Schild-Exports, Erstellungen von Warnungen, Erstellung der Log-Dateien)."},
        {"name": "--generate-emails --send-emails", "description": "Generiert und sendet die Warn-E-Mails auf Grundlage der gespeicherten Einstellungen. (Davor ist --process erforderlich.)"},
        {"name": "--send-admin-warnings", "description": "Sendet Admin-Warnungen per E-Mail an die hinterlegte Admin E-Mail-Adresse, wenn im SchildExport Klassen oder Klassenlehrkräfte vorkommen, die in den Klassen- oder Lehrkraftdaten fehlen."},
        {"name": "--no-web", "description": "Verhindert das Öffnen des Web-Interfaces."},
        {"name": "--send-log-email", "description": "Sendet eine tabellarische Übersicht (html) und die den Excel-Änderungslog (im Anhang) für einen definierten Zeitraum an die hinterlegte Admin E-Mail-Adresse."},
        {"name": "--skip-admin-warnings", "description": "Überspringt die Erstellung von Admin-Warnungen. (Darf nicht mit --send admin-warnings kombiniert werden.)"},
        {"name": "--no-log", "description": "Verhindert die Erstellung der .log Logdateien."},
        {"name": "--no-xlsx", "description": "Verhindert die Erstellung der Excel Logdateien. (Darf nicht mit --send-log-email kombiniert werden.)"},
        {"name": "--no-directory-change", "description": "Verhindert, dass Verzeichnisse über das WebEnd geändert werden können. Dazu wird der Tab in den Einstellungen entfernt und im BackEnd Funktionen blockiert."},
        {"name": "--no-directory-change --enable-upload", "description": "Verhindert, dass Verzeichnisse über das WebEnd geändert werden können und ermöglicht einen Upload von Dateien in die Verzeichnisse.\n⚠️ Aus Sicherheitsgründen sollte --enable-upload niemals ohne --no-directory-change verwendet werden!⚠️"},
        {"name": "--host", "description": "IP-Adresse, auf der der Server laufen soll (Standard: 0.0.0.0)"},
        {"name": "--port", "description": "Port, auf dem der Server laufen soll (Standard: 5000)"},
    ]
    print_success(f"Liste der verfügbaren Kommandozeilenargumente für das Befehl- und Verknüpfungs-Erstelltool wurde erstellt.")
    return jsonify({"success": True, "arguments": arguments})

# Route und Funktion zum Abrufen des Pfads der ausführbaren Datei
@app.route('/get-executable-path', methods=['GET'])
def get_executable_path():
    print_info(f"Bestimme den Pfad der ausführbaren Datei für das Befehl- und Verknüpfungs-Erstelltool...")
    # Bestimmen des Pfads der ausführbaren Datei
    if getattr(sys, 'frozen', False):
        # Wenn die Anwendung eingefroren ist (in eine .exe kompiliert mit PyInstaller)
        exe_path = sys.executable
        print_info(f"Pfad der ausführbaren Datei: {exe_path}")
    else:
        # Wenn die Anwendung in einer Standard-Python-Umgebung läuft
        exe_path = os.path.abspath(sys.argv[0])
        print_info(f"Anwendung läuft in der Python-Umgebung. Pfad der ausführbaren Datei: {exe_path}")
    return jsonify({"exePath": exe_path})

# Route und Funktion zum Erstellen einer Desktop-Verknüpfung für die Anwendung mit den angegebenen Argumenten
@app.route('/create-shortcut', methods=['POST'])
def create_shortcut():
    print_info("Erstelle Desktop-Verknüpfung...")
    data = request.json
    exe_path = data.get('exePath')
    args = data.get('args', '')

    # Empfangenen Pfad und Argumente ausgeben
    print_info(f"Empfangener exe_path: {exe_path}")
    print_info(f"Empfangene Argumente: {args}")
    print_info(f"Existiert exe_path? {os.path.exists(exe_path)}")

    if not exe_path or not os.path.exists(exe_path):
        error_message = f"Pfad der ausführbaren Datei nicht gefunden: {exe_path}"
        print_error(f"Fehler: {error_message}")
        return jsonify({"success": False, "error": error_message})

    try:
        # COM-Bibliothek initialisieren
        import pythoncom
        pythoncom.CoInitialize()  # Initialize COM
        print_info("COM-Bibliothek initialisiert.")
        
        # Desktop-Pfad abrufen
        import winshell
        desktop = winshell.desktop()
        print_info(f"Desktop-Pfad: {desktop}")

        # Pfad für die Verknüpfung festlegen
        shortcut_path = os.path.join(desktop, "Schild-WebUntis-Tool.lnk")
        print_info(f"Verknüpfungspfad: {shortcut_path}")

        # Verknüpfung erstellen
        with winshell.shortcut(shortcut_path) as shortcut:
            shortcut.path = exe_path
            shortcut.arguments = args
            shortcut.description = "Shortcut for Schild-WebUntis-Tool"
        print_success("Verknüpfung erfolgreich erstellt.")
        return jsonify({"success": True})
    except Exception as e:
        error_message = f"Fehler beim Erstellen der Verknüpfung: {str(e)}"
        print_error(error_message)
        return jsonify({"success": False, "error": error_message})
    finally:
        # COM-Bibliothek deinitialisieren
        pythoncom.CoUninitialize()
        print_info("COM-Bibliothek deinitialisiert.")

# Route und Funktion zum Hochladen von Dateien in die Quelldaten-Verzeichnisse 
@app.route('/upload-files', methods=['POST'])
def upload_files():
    print_info("Empfange Dateien zum Hochladen...")
    # Get directories from settings.ini
    config = configparser.ConfigParser()
    config.read("settings.ini", encoding='utf-8-sig')
    classes_dir = config.get("Directories", "classes_directory", fallback="./Klassendaten")
    teachers_dir = config.get("Directories", "teachers_directory", fallback="./Lehrerdaten")
    schildexport_dir = config.get("Directories", "schildexport_directory", fallback=".")

    # Create directories if they do not exist
    os.makedirs(classes_dir, exist_ok=True)
    os.makedirs(teachers_dir, exist_ok=True)
    os.makedirs(schildexport_dir, exist_ok=True)

    # Process uploaded files
    uploaded_files = request.files
    success_messages = []
    error_messages = []

    # Handle class data files
    class_data_files = request.files.getlist("class_data_files")
    for file in class_data_files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            if validate_class_data_file(file):
                # File is valid, save it
                save_path = os.path.join(classes_dir, filename)
                file.stream.seek(0)  # Ensure stream is at the beginning before saving
                file.save(save_path)
                success_messages.append(f"'{filename}' in Klassendaten hochgeladen.")
                print_info(f"Datei '{filename}' in '{classes_dir}' gespeichert.")
            else:
                error_messages.append(f"Ungültiges Dateiformat oder Inhalt: {file.filename}")
                print_warning(f"Ungültiges Dateiformat oder Inhalt: {file.filename}")
        else:
            error_messages.append(f"Ungültige Datei: {file.filename}")
            print_warning(f"Ungültige Datei: {file.filename}")

    # Handle teacher data files
    teacher_data_files = request.files.getlist("teacher_data_files")
    for file in teacher_data_files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            if validate_teacher_data_file(file):
                # File is valid, save it
                save_path = os.path.join(teachers_dir, filename)
                file.stream.seek(0)  # Ensure stream is at the beginning before saving
                file.save(save_path)
                success_messages.append(f"'{filename}' in Lehrerdaten hochgeladen.")
                print_info(f"Datei '{filename}' in '{teachers_dir}' gespeichert.")
            else:
                error_messages.append(f"Ungültiges Dateiformat oder Inhalt: {file.filename}")
                print_warning(f"Ungültiges Dateiformat oder Inhalt: {file.filename}")
        else:
            error_messages.append(f"Ungültige Datei: {file.filename}")
            print_warning(f"Ungültige Datei: {file.filename}")

    # Schild-Export Dateien (falls notwendig, hier ohne zusätzliche Validierung)
    schild_export_files = request.files.getlist("schild_export_files")
    for file in schild_export_files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            save_path = os.path.join(schildexport_dir, filename)
            file.save(save_path)
            success_messages.append(f"'{filename}' in Schild-Export hochgeladen.")
            print_info(f"Datei '{filename}' in '{schildexport_dir}' gespeichert.")
        else:
            error_messages.append(f"Ungültige Datei: {file.filename}")
            print_warning(f"Ungültige Datei: {file.filename}")

    # Prepare response message
    message = ""
    if success_messages:
        message += "Erfolgreich hochgeladen:\n" + "\n".join(success_messages)
    if error_messages:
        message += "\nFehler beim Hochladen:\n" + "\n".join(error_messages)

    return jsonify({"message": message})
# Funktion zum determinieren, welche Dateiformate erlaubt sind
def allowed_file(filename):
    # Allow only certain file extensions
    allowed_extensions = {'csv'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
#Funktion zum Überprüfen der Hochgeladenden Klassendatei auf das Vorhandensein aller erwarteter Spalten und des korrekten Seperators (;)
def validate_class_data_file(file):
    expected_columns = ["Auswahl", "", "Klasse", "Langname", "Alias", "Jahrgangsstufe", "Text", "Klassenlehrkraft", "Klassenlehrkraft", "Abteilung", "Von", "Bis"]
    expected_separator = ';'
    possible_separators = [',', ';', '\t']

    try:
        # Zurück zum Anfang der Datei
        file.stream.seek(0)
        # Erste Zeile (Header) einlesen
        first_line = file.stream.readline().decode('utf-8-sig').strip()

        # Zuerst mit dem erwarteten Separator testen
        headers = first_line.split(expected_separator)
        if headers == expected_columns:
            # Validierung erfolgreich
            file.stream.seek(0)  # Stream-Position zurücksetzen
            return True
        else:
            # Prüfen, ob ein anderer Separator das Problem verursacht
            for sep in possible_separators:
                if sep == expected_separator:
                    continue  # Bereits getestet
                headers = first_line.split(sep)
                if headers == expected_columns:
                    # Falscher Separator gefunden
                    print_warning(f"Falscher Separator in der Klassendatei '{file.filename}'. Erwartet: '{expected_separator}', gefunden: '{sep}'")
                    return False
            # Wenn kein passender Separator gefunden wurde
            print_warning(f"Ungültige Spalten oder Separator in der Klassendatei '{file.filename}'. Erwartete Spalten: {expected_columns}")
            return False
    except Exception as e:
        print_warning(f"Fehler beim Validieren der Klassendatei '{file.filename}': {e}")
        return False
#Funktion zum Überprüfen der Hochgeladenden Lehrkräftedatei auf das Vorhandensein aller erwarteter Spalten und des korrekten Seperators (Tab)
def validate_teacher_data_file(file):
    expected_columns = ["name", "longName", "foreName", "title", "birthDate", "pnr", "address.email", "address.phone", "address.mobile", "address.street", "address.postCode", "address.city"]
    expected_separator = '\t'
    possible_separators = [',', ';', '\t']

    try:
        file.stream.seek(0)
        first_line = file.stream.readline().decode('utf-8-sig').strip()

        headers = first_line.split(expected_separator)
        if headers == expected_columns:
            file.stream.seek(0)
            return True
        else:
            # Prüfen, ob ein anderer Separator das Problem verursacht
            for sep in possible_separators:
                if sep == expected_separator:
                    continue  # Bereits getestet
                headers = first_line.split(sep)
                if headers == expected_columns:
                    # Falscher Separator gefunden
                    print_warning(f"Falscher Separator in der Lehrerdaten-Datei '{file.filename}'. Erwartet: '{expected_separator}', gefunden: '{sep}'")
                    return False
            # Wenn kein passender Separator gefunden wurde
            print_warning(f"Ungültige Spalten oder Separator in der Lehrerdaten-Datei '{file.filename}'. Erwartete Spalten: {expected_columns}")
            return False
    except Exception as e:
        print_warning(f"Fehler beim Validieren der Lehrerdaten-Datei '{file.filename}': {e}")
        return False





# Funktion zum Öffnen des Standardbrowsers, automatisch auf die lokale URL, wird beim Start des Servers ausgeführt sofern nicht mit --no-web unterbunden.
def open_browser():
    host = cli_args.get("host", "127.0.0.1")
    port = cli_args.get("port", 5000)
    # Wenn der Host '0.0.0.0' ist, setzen wir ihn auf '127.0.0.1' für die Browser-URL
    if host == '0.0.0.0':
        display_host = '127.0.0.1'
    else:
        display_host = host
    url = f"http://{display_host}:{port}/"
    webbrowser.open_new(url)

# Globale Variable für CLI-Argumente
cli_args = {}

if __name__ == "__main__":
    # Banner beim Start ausgeben
    print_banner()

    ensure_ini_files_exist()

    # Parser für Kommandozeilenargumente
    parser = argparse.ArgumentParser(description="Command-line interface for processing data.")
    parser.add_argument('--process', action='store_true', help="Führt den Hauptprozess aus (Verarbeitung des Schild-Exports, Erstellungen von Warnungen, Erstellung der Log-Dateien).")
    parser.add_argument('--generate-emails', action='store_true', help="Generiert die Warn-E-Mails auf Grundlage der gespeicherten Einstellungen. (Davor ist --process erforderlich.)")
    parser.add_argument('--send-emails', action='store_true', help="Sendet die Warn-E-Mails auf Grundlage der gespeicherten Einstellungen. (Davor sind --process und --generate-emails erforderlich.)")
    parser.add_argument('--send-admin-warnings', action='store_true', help="Sendet Admin-Warnungen per E-Mail an die hinterlegte Admin E-Mail-Adresse, wenn im SchildExport Klassen oder Klassenlehrkräfte vorkommen, die in den Klassen- oder Lehrkraftdaten fehlen.")
    parser.add_argument('--no-web', action='store_true', help="Verhindert das Öffnen des Web-Interfaces.")
    parser.add_argument('--skip-admin-warnings', action='store_true', help="Überspringt die Erstellung von Admin-Warnungen. (Darf nicht mit --send admin-warnings kombiniert werden.)")
    parser.add_argument('--class-change-recipients', type=str, choices=['old', 'new', 'both'], help="Legt den Empfänger der Klassenwechsel-Warnung fest (old, new oder both). Überschreibt die ini-Konfiguration.")
    parser.add_argument('--no-log', action='store_true', help="Verhindert die Erstellung der .log Logdateien.")
    parser.add_argument('--no-xlsx', action='store_true', help="Verhindert die Erstellung der Excel Logdateien. (Darf nicht mit --send-log-email kombiniert werden.)")
    parser.add_argument('--send-log-email', action='store_true', help="Sendet eine tabellarische Übersicht (html) und die den Excel-Änderungslog (im Anhang) für einen definierten Zeitraum an die hinterlegte Admin E-Mail-Adresse.")
    parser.add_argument('--no-directory-change', action='store_true', help="Verhindert, dass Verzeichnisse über das WebEnd geändert werden können. Dazu wird der Tab in den Einstellungen entfernt und im BackEnd Funktionen blockiert.")
    parser.add_argument('--enable-upload', action='store_true', help="Ermöglicht einen Upload von Dateien in die Verzeichnisse.\n⚠️ Aus Sicherheitsgründen sollte --enable-upload niemals ohne --no-directory-change verwendet werden!⚠️")
    parser.add_argument('--host', type=str, default='0.0.0.0', help="IP-Adresse, auf der der Server laufen soll (Standard: 0.0.0.0)")
    parser.add_argument('--port', type=int, default=5000, help="Port, auf dem der Server laufen soll (Standard: 5000)")

    args = parser.parse_args()

    # Speichern der CLI-Argumente in der globalen Variable
    cli_args = {
        "process": args.process,
        "generate_emails": args.generate_emails,
        "send_emails": args.send_emails,
        "send_admin_warnings": args.send_admin_warnings,
        "no_web": args.no_web,
        "skip_admin_warnings": args.skip_admin_warnings,
        "class_change_recipients": args.class_change_recipients,
        "no_log": args.no_log,
        "no_xlsx": args.no_xlsx,
        "send_log_email": args.send_log_email,
        "no_directory_change": args.no_directory_change,
        "enable_upload": args.enable_upload,
        "host": args.host,
        "port": args.port,
    }

    # Initialisiere globale Caches
    warnings_cache = []
    generated_emails_cache = []

    # Admin-Warnungen erstellen, falls nicht übersprungen
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':  # Verhindert doppelte Ausführung im Debug-Modus
        if not args.skip_admin_warnings:
            admin_warnings()

    # Senden von Admin-Warnugen
    if args.send_admin_warnings:
        admin_warnings(send_email_flag=True)

    # Verarbeitung starten
    if args.process:
        print_info("Starte Datenverarbeitung...")
        # Übergabe der Argumente no_log und no_xlsx an process_data
        warnings_cache = process_data(
            no_log=args.no_log,
            no_xlsx=args.no_xlsx
        )
        print_success("Datenverarbeitung erfolgreich abgeschlossen.")

    # Timeframe-Import vergleichen und ggf. E-Mail versenden
    if args.send_log_email:
        print_info("Vergleiche Import-Dateien basierend auf dem definierten Zeitrahmen...")
        compare_timeframe_imports(
            no_log=args.no_log,
            no_xlsx=args.no_xlsx
        )
        print_success("Vergleich abgeschlossen.")

    # E-Mails generieren
    if args.generate_emails:
        if not warnings_cache:
            print_info("Es sind zum Generieren von E-Mails keine Warnungen im Cache vorhanden.")
        else:
            print_info("Generiere E-Mails...")
            from flask import current_app
            with app.app_context():
                generate_emails()
            print_success("E-Mails erfolgreich generiert.")

    # E-Mails senden
    if args.send_emails:
        if not generated_emails_cache:
            print_info("Es sind keine generierten E-Mails zum Senden im Cache vorhanden.")
        else:
            print_info("Sende E-Mails...")
            for email in generated_emails_cache:
                send_email(email['subject'], email['body'], email['to'])
            print_success("E-Mails erfolgreich gesendet.")


    # Prüfen, ob eine rein aufgabenspezifische CLI-Zugehörigkeit vorliegt und der Server übergangen werden soll.
    if any([args.process, args.generate_emails, args.send_emails, args.send_admin_warnings, args.send_log_email]):
        print_success("CLI-Verarbeitung abgeschlossen. Beende das Skript.")
        sys.exit(0)

    # Starten des WSGI-Servers mit Waitress
    if not args.no_web:
        browser_thread = threading.Thread(target=open_browser)
        browser_thread.start()

    try:
        # Initialer Check der Historie
        import sqlite3
        import history_manager
        import configparser
        conn = sqlite3.connect('history.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='comparisons'")
        if cursor.fetchone():
            cursor.execute("SELECT COUNT(*) FROM comparisons")
            count = cursor.fetchone()[0]
            if count == 0:
                print_info("✨ Historie ist leer. Starte automatische Indizierung vorhandener Logs...")
                config = configparser.ConfigParser()
                from utils import safe_read_config
                safe_read_config(config, 'settings.ini')
                log_dir = config.get("Directories", "log_directory", fallback="./Logs")
                reindexed = history_manager.reindex_logs(log_dir)
                if reindexed > 0:
                    print_success(f"✅ {reindexed} historische Log-Dateien wurden automatisch indiziert.")
            else:
                print_success(f"📦 Historien-Datenbank bereit ({count} Vergleiche geladen).")
        conn.close()
    except Exception as e:
        print_warning(f"Hinweis zur Historie-Initialisierung: {e}")

    try:
        serve(app, host=cli_args.get("host", "0.0.0.0"), port=cli_args.get("port", 5000))
    except Exception as e:
        print_error(f"Fehler beim Starten des WSGI-Servers: {e}")



__all__ = ['admin_warnings_cache','admin_warnings']





